"""
Export Utilities
PDF and Excel export functionality for Patent Analytics Platform
"""

import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from django.http import HttpResponse


class ExcelExporter:
    """Generate Excel files using openpyxl"""

    def __init__(self, title: str = "Export"):
        self.title = title
        self.workbook = None
        self.current_sheet = None

    def create_workbook(self):
        """Create a new workbook"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.current_sheet.title = self.title[:31]  # Excel sheet name limit

        # Store styles for reuse
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        return self

    def add_sheet(self, name: str):
        """Add a new sheet"""
        self.current_sheet = self.workbook.create_sheet(title=name[:31])
        return self

    def write_header(self, headers: List[str], row: int = 1):
        """Write header row with styling"""
        for col, header in enumerate(headers, 1):
            cell = self.current_sheet.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        return self

    def write_row(self, data: List[Any], row: int):
        """Write a data row"""
        for col, value in enumerate(data, 1):
            cell = self.current_sheet.cell(row=row, column=col, value=value)
            cell.border = self.thin_border
        return self

    def write_data(self, headers: List[str], data: List[List[Any]], start_row: int = 1):
        """Write headers and data rows"""
        self.write_header(headers, start_row)
        for row_idx, row_data in enumerate(data, start_row + 1):
            self.write_row(row_data, row_idx)
        return self

    def auto_fit_columns(self):
        """Auto-fit column widths"""
        for column in self.current_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            self.current_sheet.column_dimensions[column_letter].width = adjusted_width
        return self

    def get_response(self, filename: str = "export.xlsx") -> HttpResponse:
        """Return HTTP response with Excel file"""
        self.auto_fit_columns()

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PDFExporter:
    """Generate PDF files using reportlab"""

    def __init__(self, title: str = "Export", orientation: str = "portrait"):
        self.title = title
        self.orientation = orientation
        self.buffer = io.BytesIO()
        self.elements = []
        self.styles = None

    def create_document(self):
        """Initialize PDF document"""
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib import colors

        self.styles = getSampleStyleSheet()

        # Add custom styles
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            alignment=TA_CENTER,
            spaceAfter=30,
            textColor=colors.HexColor('#000000')
        ))

        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#00D9FF')
        ))

        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.white
        ))

        self.page_size = landscape(letter) if self.orientation == "landscape" else letter
        return self

    def add_title(self, text: str):
        """Add title to document"""
        from reportlab.platypus import Paragraph, Spacer

        self.elements.append(Paragraph(text, self.styles['ReportTitle']))
        self.elements.append(Spacer(1, 12))
        return self

    def add_section(self, title: str):
        """Add section header"""
        from reportlab.platypus import Paragraph

        self.elements.append(Paragraph(title, self.styles['SectionHeader']))
        return self

    def add_paragraph(self, text: str):
        """Add paragraph text"""
        from reportlab.platypus import Paragraph, Spacer

        self.elements.append(Paragraph(text, self.styles['Normal']))
        self.elements.append(Spacer(1, 6))
        return self

    def add_table(self, headers: List[str], data: List[List[Any]], col_widths: Optional[List[int]] = None):
        """Add table to document"""
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib import colors

        table_data = [headers] + data

        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00D9FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        self.elements.append(table)
        self.elements.append(Spacer(1, 20))
        return self

    def add_spacer(self, height: int = 12):
        """Add vertical space"""
        from reportlab.platypus import Spacer

        self.elements.append(Spacer(1, height))
        return self

    def add_page_break(self):
        """Add page break"""
        from reportlab.platypus import PageBreak

        self.elements.append(PageBreak())
        return self

    def get_response(self, filename: str = "export.pdf") -> HttpResponse:
        """Return HTTP response with PDF file"""
        from reportlab.platypus import SimpleDocTemplate

        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.page_size,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )

        doc.build(self.elements)
        self.buffer.seek(0)

        response = HttpResponse(
            self.buffer.getvalue(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def export_infringement_report_pdf(analysis_data: Dict[str, Any]) -> HttpResponse:
    """Generate infringement analysis PDF report"""
    exporter = PDFExporter(title="Infringement Analysis Report", orientation="portrait")
    exporter.create_document()

    # Title
    exporter.add_title("Patent Infringement Analysis Report")

    # Summary section
    exporter.add_section("Executive Summary")
    exporter.add_paragraph(f"Target Patent: {analysis_data.get('target_patent', 'N/A')}")
    exporter.add_paragraph(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d')}")
    exporter.add_paragraph(f"Overall Risk Level: {analysis_data.get('risk_level', 'N/A')}")

    # Claims analysis
    claims = analysis_data.get('claims', [])
    if claims:
        exporter.add_section("Claims Analysis")
        headers = ['Claim #', 'Type', 'Risk Level', 'Coverage %', 'Notes']
        data = [
            [
                str(c.get('claim_number', '')),
                c.get('claim_type', ''),
                c.get('risk_level', ''),
                f"{c.get('coverage_percentage', 0)}%",
                c.get('notes', '')[:50]
            ]
            for c in claims
        ]
        exporter.add_table(headers, data)

    # Evidence section
    evidence = analysis_data.get('evidence', [])
    if evidence:
        exporter.add_section("Supporting Evidence")
        headers = ['Reference', 'Type', 'Relevance', 'Summary']
        data = [
            [
                e.get('reference_id', ''),
                e.get('type', ''),
                e.get('relevance', ''),
                e.get('summary', '')[:60]
            ]
            for e in evidence
        ]
        exporter.add_table(headers, data)

    # Recommendations
    exporter.add_section("Recommendations")
    for rec in analysis_data.get('recommendations', []):
        exporter.add_paragraph(f"- {rec}")

    filename = f"infringement_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return exporter.get_response(filename)


def export_infringement_report_excel(analysis_data: Dict[str, Any]) -> HttpResponse:
    """Generate infringement analysis Excel report"""
    exporter = ExcelExporter(title="Infringement Analysis")
    exporter.create_workbook()

    # Summary sheet
    exporter.write_data(
        headers=['Field', 'Value'],
        data=[
            ['Target Patent', analysis_data.get('target_patent', 'N/A')],
            ['Analysis Date', datetime.now().strftime('%Y-%m-%d')],
            ['Risk Level', analysis_data.get('risk_level', 'N/A')],
            ['Total Claims', str(len(analysis_data.get('claims', [])))],
        ]
    )

    # Claims sheet
    claims = analysis_data.get('claims', [])
    if claims:
        exporter.add_sheet('Claims Analysis')
        headers = ['Claim #', 'Type', 'Risk Level', 'Coverage %', 'Notes']
        data = [
            [
                c.get('claim_number', ''),
                c.get('claim_type', ''),
                c.get('risk_level', ''),
                c.get('coverage_percentage', 0),
                c.get('notes', '')
            ]
            for c in claims
        ]
        exporter.write_data(headers, data)

    # Evidence sheet
    evidence = analysis_data.get('evidence', [])
    if evidence:
        exporter.add_sheet('Evidence')
        headers = ['Reference', 'Type', 'Relevance', 'Summary']
        data = [
            [
                e.get('reference_id', ''),
                e.get('type', ''),
                e.get('relevance', ''),
                e.get('summary', '')
            ]
            for e in evidence
        ]
        exporter.write_data(headers, data)

    filename = f"infringement_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return exporter.get_response(filename)


def export_prior_art_report_pdf(report_data: Dict[str, Any]) -> HttpResponse:
    """Generate prior art search PDF report"""
    exporter = PDFExporter(title="Prior Art Report", orientation="portrait")
    exporter.create_document()

    # Title
    report_type = report_data.get('report_type', 'Prior Art Search')
    exporter.add_title(f"{report_type} Report")

    # Summary
    exporter.add_section("Project Summary")
    exporter.add_paragraph(f"Project: {report_data.get('project_name', 'N/A')}")
    exporter.add_paragraph(f"Target Patent: {report_data.get('target_patent', 'N/A')}")
    exporter.add_paragraph(f"Search Date Range: {report_data.get('date_range', 'N/A')}")
    exporter.add_paragraph(f"Total References Found: {report_data.get('total_references', 0)}")
    exporter.add_paragraph(f"Selected References: {report_data.get('selected_references', 0)}")

    # Evidence table
    evidence = report_data.get('evidence', [])
    if evidence:
        exporter.add_section("Prior Art References")
        headers = ['Reference', 'Date', 'Type', 'Relevance', 'Title']
        data = [
            [
                e.get('reference_id', ''),
                e.get('publication_date', ''),
                e.get('evidence_type', ''),
                e.get('relevance_level', ''),
                e.get('title', '')[:40]
            ]
            for e in evidence[:50]  # Limit to 50 entries
        ]
        exporter.add_table(headers, data)

    # Claim chart if available
    claim_mappings = report_data.get('claim_mappings', [])
    if claim_mappings:
        exporter.add_page_break()
        exporter.add_section("Claim Chart")
        headers = ['Claim', 'Reference', 'Coverage', 'Analysis']
        data = [
            [
                f"Claim {m.get('claim_number', '')}",
                m.get('reference_id', ''),
                f"{m.get('coverage_percentage', 0)}%",
                m.get('analysis', '')[:50]
            ]
            for m in claim_mappings
        ]
        exporter.add_table(headers, data)

    filename = f"prior_art_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return exporter.get_response(filename)


def export_prior_art_report_excel(report_data: Dict[str, Any]) -> HttpResponse:
    """Generate prior art search Excel report"""
    exporter = ExcelExporter(title="Prior Art Report")
    exporter.create_workbook()

    # Summary sheet
    exporter.write_data(
        headers=['Field', 'Value'],
        data=[
            ['Project', report_data.get('project_name', 'N/A')],
            ['Report Type', report_data.get('report_type', 'N/A')],
            ['Target Patent', report_data.get('target_patent', 'N/A')],
            ['Search Date Range', report_data.get('date_range', 'N/A')],
            ['Total References', str(report_data.get('total_references', 0))],
            ['Selected References', str(report_data.get('selected_references', 0))],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
    )

    # Evidence sheet
    evidence = report_data.get('evidence', [])
    if evidence:
        exporter.add_sheet('Prior Art References')
        headers = ['Reference ID', 'Publication Date', 'Type', 'Relevance', 'Score', 'Title', 'Summary']
        data = [
            [
                e.get('reference_id', ''),
                e.get('publication_date', ''),
                e.get('evidence_type', ''),
                e.get('relevance_level', ''),
                e.get('relevance_score', ''),
                e.get('title', ''),
                e.get('summary', '')
            ]
            for e in evidence
        ]
        exporter.write_data(headers, data)

    # Claim mappings sheet
    claim_mappings = report_data.get('claim_mappings', [])
    if claim_mappings:
        exporter.add_sheet('Claim Chart')
        headers = ['Claim #', 'Reference', 'Coverage Type', 'Coverage %', 'Confidence', 'Analysis']
        data = [
            [
                m.get('claim_number', ''),
                m.get('reference_id', ''),
                m.get('coverage_type', ''),
                m.get('coverage_percentage', ''),
                m.get('confidence', ''),
                m.get('analysis', '')
            ]
            for m in claim_mappings
        ]
        exporter.write_data(headers, data)

    filename = f"prior_art_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return exporter.get_response(filename)


def export_claim_chart_excel(chart_data: Dict[str, Any]) -> HttpResponse:
    """Generate claim chart Excel export"""
    exporter = ExcelExporter(title="Claim Chart")
    exporter.create_workbook()

    # Header info
    exporter.write_data(
        headers=['Field', 'Value'],
        data=[
            ['Target Patent', chart_data.get('target_patent', 'N/A')],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
    )

    # Main claim chart
    claims = chart_data.get('claims', [])
    if claims:
        exporter.add_sheet('Claim Chart')
        headers = ['Claim #', 'Claim Text', 'Reference', 'Relevant Disclosure', 'Coverage', 'Notes']
        data = []
        for claim in claims:
            for mapping in claim.get('mappings', []):
                data.append([
                    claim.get('claim_number', ''),
                    claim.get('claim_text', '')[:100],
                    mapping.get('reference_id', ''),
                    mapping.get('disclosure', ''),
                    f"{mapping.get('coverage_percentage', 0)}%",
                    mapping.get('notes', '')
                ])
        exporter.write_data(headers, data)

    filename = f"claim_chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return exporter.get_response(filename)
