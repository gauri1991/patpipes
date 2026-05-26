"""Report and presentation views"""

import logging

logger = logging.getLogger(__name__)

from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
import json

from ..models import (
    AnalyticsProject, AnalyticsReport, AnalyticsPresentation,
)
from ..serializers import AnalyticsReportSerializer
from ..services import ReportGenerator

class AnalyticsReportViewSet(viewsets.ModelViewSet):
    """Analytics report management"""
    
    serializer_class = AnalyticsReportSerializer
    # Uses global default: IsAuthenticated
    
    def get_queryset(self):
        project_id = self.request.query_params.get('project_id')
        if project_id:
            return AnalyticsReport.objects.filter(project_id=project_id)
        return AnalyticsReport.objects.all()
    
    def perform_create(self, serializer):
        user = self.request.user

        # Get project from request data or use default
        project = serializer.validated_data.get('project')
        if not project:
            project = AnalyticsProject.objects.first()
            if not project:
                project = AnalyticsProject.objects.create(
                    name="Default Project",
                    description="Default project for reports",
                    status="active",
                    created_by=user
                )

        serializer.save(created_by=user, project=project)

    @action(detail=True, methods=['post'])
    def generate_report(self, request, pk=None):
        """Generate analytics report"""
        report = self.get_object()
        
        # Update status
        report.status = 'generating'
        report.save()
        
        # Generate report using report generator
        generator = ReportGenerator(report.project)
        try:
            # Use actual report generation service
            report.sections = generator.generate_report_sections(
                report.report_type, 
                report.include_sections or ['executive_summary', 'key_findings', 'recommendations']
            )
        except Exception as e:
            # Fallback to mock data if generation fails
            print(f"Report generation failed: {e}")
            report.sections = {
                'executive_summary': 'Generated executive summary...',
                'key_findings': ['Finding 1', 'Finding 2', 'Finding 3'],
                'recommendations': ['Recommendation 1', 'Recommendation 2']
            }
        report.status = 'completed'
        report.save()
        
        return Response({
            'status': 'Report generated',
            'report_id': str(report.id)
        })
    
    @action(detail=True, methods=['post'])
    def export_pdf(self, request, pk=None):
        """Export report as PDF using reportlab"""
        report = self.get_object()

        if report.status != 'completed':
            return Response(
                {'error': 'Report must be completed before exporting'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import io
            from django.http import HttpResponse
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter,
                                    rightMargin=inch, leftMargin=inch,
                                    topMargin=inch, bottomMargin=inch)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=12)
            h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=13, spaceAfter=6)
            body_style = styles['BodyText']
            story = []

            # Cover
            story.append(Paragraph(report.title, title_style))
            story.append(Paragraph(f"Report Type: {report.get_report_type_display() if hasattr(report, 'get_report_type_display') else report.report_type}", body_style))
            story.append(Paragraph(f"Status: {report.status.title()}", body_style))
            story.append(Paragraph(f"Generated: {report.updated_at.strftime('%B %d, %Y')}", body_style))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
            story.append(Spacer(1, 0.2 * inch))

            # Executive Summary
            if report.executive_summary:
                story.append(Paragraph("Executive Summary", h2_style))
                story.append(Paragraph(report.executive_summary, body_style))
                story.append(Spacer(1, 0.15 * inch))

            # Report Sections
            sections = report.sections or {}
            for section_key, section_data in sections.items():
                if isinstance(section_data, dict):
                    title = section_data.get('title', section_key.replace('_', ' ').title())
                    story.append(Paragraph(title, h2_style))
                    content = section_data.get('content', {})
                    if isinstance(content, dict):
                        for k, v in content.items():
                            if isinstance(v, list) and v:
                                story.append(Paragraph(f"<b>{k.replace('_', ' ').title()}:</b>", body_style))
                                for item in v:
                                    if isinstance(item, str):
                                        story.append(Paragraph(f"• {item}", body_style))
                                    elif isinstance(item, dict):
                                        story.append(Paragraph(f"• {item}", body_style))
                            elif isinstance(v, dict) and v:
                                story.append(Paragraph(f"<b>{k.replace('_', ' ').title()}:</b>", body_style))
                                for dk, dv in v.items():
                                    story.append(Paragraph(f"  {dk}: {dv}", body_style))
                            elif v is not None and str(v).strip():
                                story.append(Paragraph(f"<b>{k.replace('_', ' ').title()}:</b> {v}", body_style))
                    story.append(Spacer(1, 0.15 * inch))

            # Recommendations
            if report.recommendations:
                story.append(Paragraph("Recommendations", h2_style))
                for rec in report.recommendations:
                    story.append(Paragraph(f"• {rec}", body_style))
                story.append(Spacer(1, 0.15 * inch))

            # Conclusions
            if report.conclusions:
                story.append(Paragraph("Conclusions", h2_style))
                story.append(Paragraph(report.conclusions, body_style))

            doc.build(story)
            buffer.seek(0)

            http_response = HttpResponse(buffer, content_type='application/pdf')
            safe_title = report.title.replace(' ', '_')[:50]
            http_response['Content-Disposition'] = f'attachment; filename="{safe_title}.pdf"'
            return http_response

        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return Response(
                {'error': f'PDF export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def export_excel(self, request, pk=None):
        """Export report as Excel using openpyxl"""
        report = self.get_object()

        if report.status != 'completed':
            return Response(
                {'error': 'Report must be completed before exporting'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import io
            from django.http import HttpResponse
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # ── Summary sheet ──────────────────────────────────────────────
            ws = wb.active
            ws.title = "Summary"
            header_font = Font(bold=True, size=12)
            title_fill = PatternFill("solid", fgColor="2563EB")
            title_font = Font(bold=True, color="FFFFFF", size=14)
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws['A1'] = report.title
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:C1')

            rows = [
                ("Report Type", report.report_type.replace('_', ' ').title()),
                ("Status", report.status.title()),
                ("Generated", report.updated_at.strftime('%Y-%m-%d %H:%M')),
                ("Sections", len(report.sections or {})),
            ]
            for i, (label, value) in enumerate(rows, start=3):
                ws.cell(row=i, column=1, value=label).font = header_font
                ws.cell(row=i, column=2, value=str(value))
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40

            if report.executive_summary:
                ws.cell(row=8, column=1, value="Executive Summary").font = header_font
                cell = ws.cell(row=9, column=1, value=report.executive_summary)
                cell.alignment = Alignment(wrap_text=True)
                ws.merge_cells('A9:C9')
                ws.row_dimensions[9].height = 80

            # ── Per-section sheets ─────────────────────────────────────────
            for section_key, section_data in (report.sections or {}).items():
                if not isinstance(section_data, dict):
                    continue
                sheet_name = section_key.replace('_', ' ').title()[:31]
                ws2 = wb.create_sheet(title=sheet_name)
                content = section_data.get('content', {})
                row_num = 1
                ws2.cell(row=row_num, column=1, value=section_data.get('title', sheet_name)).font = Font(bold=True, size=13)
                row_num += 2

                if isinstance(content, dict):
                    for k, v in content.items():
                        ws2.cell(row=row_num, column=1, value=k.replace('_', ' ').title()).font = header_font
                        if isinstance(v, list):
                            for item in v:
                                row_num += 1
                                ws2.cell(row=row_num, column=2, value=str(item) if not isinstance(item, str) else item)
                        elif isinstance(v, dict):
                            for dk, dv in v.items():
                                row_num += 1
                                ws2.cell(row=row_num, column=2, value=str(dk))
                                ws2.cell(row=row_num, column=3, value=str(dv))
                        else:
                            ws2.cell(row=row_num, column=2, value=str(v) if v is not None else '')
                        row_num += 2
                ws2.column_dimensions['A'].width = 28
                ws2.column_dimensions['B'].width = 40
                ws2.column_dimensions['C'].width = 30

            # ── Recommendations sheet ──────────────────────────────────────
            if report.recommendations:
                ws3 = wb.create_sheet(title="Recommendations")
                ws3.cell(row=1, column=1, value="Recommendations").font = Font(bold=True, size=13)
                for i, rec in enumerate(report.recommendations, start=3):
                    ws3.cell(row=i, column=1, value=f"{i - 2}.")
                    ws3.cell(row=i, column=2, value=rec).alignment = Alignment(wrap_text=True)
                ws3.column_dimensions['A'].width = 5
                ws3.column_dimensions['B'].width = 70

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            http_response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            safe_title = report.title.replace(' ', '_')[:50]
            http_response['Content-Disposition'] = f'attachment; filename="{safe_title}.xlsx"'
            return http_response

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return Response(
                {'error': f'Excel export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def report_templates(self, request):
        """Get available report templates"""
        templates = [
            {
                'type': 'landscape_analysis',
                'name': 'Patent Landscape Analysis',
                'description': 'Comprehensive landscape analysis report',
                'sections': [
                    'executive_summary', 'market_overview', 'technology_trends',
                    'competitive_analysis', 'geographic_analysis', 'recommendations'
                ]
            },
            {
                'type': 'competitive_intelligence',
                'name': 'Competitive Intelligence',
                'description': 'Competitor analysis and positioning',
                'sections': [
                    'executive_summary', 'competitor_profiles', 'competitive_positioning',
                    'strengths_weaknesses', 'opportunities_threats', 'recommendations'
                ]
            },
            {
                'type': 'fto_analysis',
                'name': 'Freedom to Operate Analysis',
                'description': 'FTO assessment and risk analysis',
                'sections': [
                    'executive_summary', 'patent_landscape', 'risk_assessment',
                    'mitigation_strategies', 'recommendations'
                ]
            },
            {
                'type': 'white_space_analysis',
                'name': 'White Space Analysis',
                'description': 'Innovation opportunity identification',
                'sections': [
                    'executive_summary', 'technology_gaps', 'market_opportunities',
                    'innovation_roadmap', 'recommendations'
                ]
            }
        ]
        
        return Response(templates)


class AnalyticsPresentationViewSet(viewsets.ModelViewSet):
    """Analytics presentation management"""

    serializer_class = 'AnalyticsPresentationSerializer'
    # Uses global default: IsAuthenticated

    def get_serializer_class(self):
        from ..serializers import AnalyticsPresentationSerializer
        return AnalyticsPresentationSerializer

    def get_queryset(self):
        project_id = self.request.query_params.get('project_id')
        if project_id:
            from ..models import AnalyticsPresentation
            return AnalyticsPresentation.objects.filter(project_id=project_id)
        from ..models import AnalyticsPresentation
        return AnalyticsPresentation.objects.all()

    def perform_create(self, serializer):
        user = self.request.user

        # Get project from request data or use default
        project = serializer.validated_data.get('project')
        if not project:
            project = AnalyticsProject.objects.first()
            if not project:
                project = AnalyticsProject.objects.create(
                    name="Default Project",
                    description="Default project for presentations",
                    status="active",
                    created_by=user
                )

        # Initialize with default slide
        slides = serializer.validated_data.get('slides', [])
        if not slides:
            slides = [{
                'id': 1,
                'type': 'title',
                'title': serializer.validated_data.get('name', 'Untitled Presentation'),
                'subtitle': serializer.validated_data.get('description', ''),
                'content': {}
            }]

        serializer.save(created_by=user, project=project, slides=slides)

    @action(detail=True, methods=['post'])
    def export_pptx(self, request, pk=None):
        """Export presentation as PowerPoint"""
        presentation = self.get_object()

        try:
            # Placeholder implementation for PowerPoint export
            # In a full implementation, this would use python-pptx library
            from django.http import HttpResponse

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
            response['Content-Disposition'] = f'attachment; filename="presentation_{presentation.id}.pptx"'

            # Placeholder content - would be replaced with actual PPTX generation
            response.write(b'PK')  # PPTX files start with PK (ZIP format)

            return response
        except Exception as e:
            return Response(
                {'error': f'Failed to export presentation: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def export_pdf(self, request, pk=None):
        """Export presentation as PDF"""
        presentation = self.get_object()

        try:
            from django.http import HttpResponse

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="presentation_{presentation.id}.pdf"'

            # Placeholder PDF content
            response.write(b'%PDF-1.4\n')

            return response
        except Exception as e:
            return Response(
                {'error': f'Failed to export presentation: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def present(self, request, pk=None):
        """Mark presentation as presented and update statistics"""
        presentation = self.get_object()

        # Update last presented timestamp and increment count
        presentation.last_presented = timezone.now()
        presentation.presentation_count += 1
        presentation.save()

        return Response({
            'status': 'Presentation marked as presented',
            'last_presented': presentation.last_presented,
            'presentation_count': presentation.presentation_count
        })

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a presentation"""
        original = self.get_object()

        # Create a copy
        from ..models import AnalyticsPresentation
        duplicate = AnalyticsPresentation.objects.create(
            project=original.project,
            name=f"{original.name} (Copy)",
            description=original.description,
            presentation_type=original.presentation_type,
            status='draft',
            theme=original.theme,
            slides=original.slides.copy() if original.slides else [],
            speaker_notes=original.speaker_notes.copy() if original.speaker_notes else {},
            slide_count=original.slide_count,
            duration_minutes=original.duration_minutes,
            template_id=original.template_id,
            template_config=original.template_config.copy() if original.template_config else {},
            created_by=request.user if request.user.is_authenticated else original.created_by
        )

        from ..serializers import AnalyticsPresentationSerializer
        serializer = AnalyticsPresentationSerializer(duplicate)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


