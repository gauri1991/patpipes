"""Analytics project views"""

import logging

logger = logging.getLogger(__name__)

from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Count, Q, Avg, Sum
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from datetime import timedelta
from collections import defaultdict
import json

from ..models import (
    AnalyticsProject, TechnologyArea, PatentDataset, CompetitorProfile,
    GlobalCompetitorProfile, GlobalTechnologyArea,
    AnalyticsVisualization, AnalyticsReport, AnalyticsPresentation, AnalyticsInsight, PatentRecord,
    ColumnMappingRule, Template, TemplateUsageLog, PatentAnalysisResult, SalesPackage
)
from ..serializers import (
    AnalyticsProjectSerializer, AnalyticsProjectCreateSerializer,
    TechnologyAreaSerializer, PatentDatasetSerializer, PatentDatasetListSerializer,
    CompetitorProfileSerializer,
    GlobalCompetitorProfileSerializer, GlobalTechnologyAreaSerializer,
    AnalyticsVisualizationSerializer, AnalyticsReportSerializer,
    AnalyticsInsightSerializer, AnalyticsDashboardSerializer, SalesPackageSerializer
)
from ..serializers_simple import SimpleAnalyticsProjectSerializer
from ..services import AnalyticsDataProcessor, ReportGenerator, VisualizationEngine
from ..file_processors import process_patent_dataset

class AnalyticsProjectViewSet(viewsets.ModelViewSet):
    """Analytics project management"""

    queryset = AnalyticsProject.objects.all()
    # Uses global default: IsAuthenticated
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'updated_at', 'name', 'status', 'priority']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return AnalyticsProjectCreateSerializer
        # Use simplified serializer temporarily to avoid database issues
        return SimpleAnalyticsProjectSerializer
    
    def get_queryset(self):
        user = self.request.user
        
        # Start with base queryset based on user permissions
        base = AnalyticsProject.objects.select_related(
            'portfolio', 'created_by', 'assigned_to'
        ).annotate(
            _dataset_count=Count('datasets'),
            _viz_count=Count('visualizations'),
            _report_count=Count('reports'),
            _dataset_done=Count('datasets', filter=Q(datasets__processing_status='completed')),
            _viz_done=Count('visualizations', filter=Q(visualizations__status='completed')),
            _report_done=Count('reports', filter=Q(reports__status='completed')),
        )
        if hasattr(user, 'role') and user.role in ['admin', 'manager']:
            queryset = base.all()
        else:
            queryset = base.filter(Q(created_by=user) | Q(assigned_to=user))

        # Date range filtering
        created_after = self.request.query_params.get('created_after')
        created_before = self.request.query_params.get('created_before')
        if created_after:
            queryset = queryset.filter(created_at__gte=created_after)
        if created_before:
            queryset = queryset.filter(created_at__lte=created_before)

        return queryset.order_by('-updated_at')

    def perform_create(self, serializer):
        """Handle project creation"""
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get analytics dashboard data — cached per-user for 5 minutes."""
        from django.core.cache import cache
        from django.db.models import Sum
        from django.db.models.functions import TruncMonth, Substr
        from django.db.models.expressions import Value
        from datetime import datetime, timedelta

        user = request.user
        cache_key = f'analytics_dashboard_{user.pk}'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        projects = self.get_queryset()
        project_ids = list(projects.values_list('id', flat=True))

        # ── Core counts (single aggregation pass) ──
        overview = projects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status__in=[
                'active', 'scope_definition', 'data_collection',
                'patent_analysis', 'visualization', 'report_generation'
            ])),
            completed=Count('id', filter=Q(status='completed')),
        )
        total_projects = overview['total']
        active_projects = overview['active']
        completed_projects = overview['completed']

        # Dataset and patent metrics
        datasets = PatentDataset.objects.filter(project__in=project_ids)
        total_datasets = datasets.count()
        total_patents_in_datasets = datasets.aggregate(total=Sum('total_patents'))['total'] or 0

        from domains.patents.models import Patent
        portfolio_ids = list(
            projects.exclude(portfolio=None).values_list('portfolio_id', flat=True)
        )
        total_patents_in_portfolios = (
            Patent.objects.filter(portfolio__in=portfolio_ids).count()
            if portfolio_ids else 0
        )

        patents_with_ai_analysis = (
            PatentAnalysisResult.objects.values('application_id').distinct().count()
        )
        total_visualizations = (
            AnalyticsVisualization.objects.filter(project__in=project_ids).count()
        )

        # Recent activity (small, safe queries)
        recent_projects = projects.order_by('-updated_at')[:5]
        recent_insights = list(
            AnalyticsInsight.objects.filter(project__in=project_ids).order_by('-created_at')[:10]
        )

        # Statistics
        projects_by_status = dict(
            projects.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )

        # Technology areas distribution — pure SQL aggregation, no Python loops
        technology_areas_distribution = dict(
            TechnologyArea.objects.filter(project__in=project_ids)
            .values('name').annotate(count=Count('id')).values_list('name', 'count')
        )

        # Monthly project trends (last 12 months)
        twelve_months_ago = datetime.now() - timedelta(days=365)
        monthly_trends_qs = (
            projects.filter(created_at__gte=twelve_months_ago)
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(projects=Count('id'), completed=Count('id', filter=Q(status='completed')))
            .order_by('month')
        )
        monthly_project_trends = []
        completion_rate_trend = []
        for trend in monthly_trends_qs:
            month_str = trend['month'].strftime('%Y-%m') if trend['month'] else 'Unknown'
            monthly_project_trends.append({'month': month_str, 'projects': trend['projects']})
            total_m = trend['projects']
            rate = round((trend['completed'] / total_m * 100) if total_m > 0 else 0, 1)
            completion_rate_trend.append({'month': month_str, 'rate': rate})

        if not monthly_project_trends:
            current_month = datetime.now().strftime('%Y-%m')
            monthly_project_trends = [{'month': current_month, 'projects': total_projects}]
        if not completion_rate_trend:
            current_rate = round((completed_projects / total_projects * 100) if total_projects > 0 else 0, 1)
            current_month = datetime.now().strftime('%Y-%m')
            completion_rate_trend = [{'month': current_month, 'rate': current_rate}]

        # ── Patent-level KPIs — DB aggregation, NO Python loops ──
        all_records = PatentRecord.objects.filter(dataset__project__in=project_ids)

        # Top IPC codes from PatentRecord (first 4 chars = subclass-level key)
        # Using DB Substr so no Python iteration
        top_ipc_qs = (
            all_records
            .exclude(ipc_classification='')
            .exclude(ipc_classification__isnull=True)
            .annotate(code=Substr('ipc_classification', 1, 4))
            .values('code')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )
        top_technology_areas = [{'code': r['code'], 'count': r['count']} for r in top_ipc_qs]

        # Top assignees from PatentRecord — pure SQL GROUP BY
        top_assignees_qs = (
            all_records
            .exclude(assignee='')
            .exclude(assignee__isnull=True)
            .values('assignee')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )
        top_assignees = [{'name': r['assignee'], 'count': r['count']} for r in top_assignees_qs]

        # Patent status distribution — pure SQL
        patent_status_dist = {}
        if portfolio_ids:
            patent_status_dist = dict(
                Patent.objects.filter(portfolio__in=portfolio_ids)
                .values('status').annotate(count=Count('id')).values_list('status', 'count')
            )

        # Filing trend by year — pure SQL GROUP BY year
        filing_trend_qs = (
            all_records
            .exclude(filing_date__isnull=True)
            .filter(filing_date__year__gte=2010)
            .values('filing_date__year')
            .annotate(count=Count('id'))
            .order_by('filing_date__year')
        )
        filing_trend_data = [
            {'year': r['filing_date__year'], 'count': r['count']} for r in filing_trend_qs
        ]

        dashboard_data = {
            'total_projects': total_projects,
            'active_projects': active_projects,
            'completed_projects': completed_projects,
            'total_datasets': total_datasets,
            'total_patents_analyzed': total_patents_in_datasets,
            'total_patents_in_portfolios': total_patents_in_portfolios,
            'patents_with_ai_analysis': patents_with_ai_analysis,
            'total_visualizations': total_visualizations,
            'recent_projects': recent_projects,
            'recent_insights': recent_insights,
            'projects_by_status': projects_by_status,
            'projects_by_type': dict(
                projects.values('analysis_scope__type').annotate(count=Count('id')).values_list('analysis_scope__type', 'count')
            ) or {'No Type Specified': total_projects},
            'technology_areas_distribution': technology_areas_distribution,
            'monthly_project_trends': monthly_project_trends,
            'completion_rate_trend': completion_rate_trend,
            'top_technology_areas': top_technology_areas,
            'top_assignees': top_assignees,
            'patent_status_distribution': patent_status_dist,
            'filing_trend': filing_trend_data,
        }

        serializer = AnalyticsDashboardSerializer(dashboard_data)
        result = serializer.data
        cache.set(cache_key, result, timeout=300)  # 5-minute cache
        return Response(result)

    @action(detail=False, methods=['post'])
    def compare_projects(self, request):
        """Compare 2+ projects side by side."""
        project_ids = request.data.get('project_ids', [])
        if len(project_ids) < 2:
            return Response({'error': 'At least 2 project IDs required'}, status=status.HTTP_400_BAD_REQUEST)

        from ..patent_data_service import get_project_patent_count

        comparisons = []
        for pid in project_ids[:5]:  # max 5
            try:
                project = AnalyticsProject.objects.get(id=pid)
            except AnalyticsProject.DoesNotExist:
                continue

            patent_count = get_project_patent_count(str(pid))
            datasets_count = PatentDataset.objects.filter(project=project).count()
            viz_count = AnalyticsVisualization.objects.filter(project=project).count()
            reports_count = AnalyticsReport.objects.filter(project=project).count()

            comparisons.append({
                'project_id': str(project.id),
                'name': project.name,
                'status': project.status,
                'type': project.analysis_scope.get('type', '') if project.analysis_scope else '',
                'patent_count': patent_count,
                'datasets_count': datasets_count,
                'visualizations_count': viz_count,
                'reports_count': reports_count,
                'created_at': project.created_at.isoformat() if project.created_at else None,
                'progress': project.progress_percentage,
            })

        return Response({'comparisons': comparisons})

    @action(detail=True, methods=['post'])
    def phase_recommendations(self, request, pk=None):
        """Generate AI recommendations for a project phase using the configured LLM provider."""
        project = self.get_object()
        phase_name = request.data.get('phase_name', '')
        phase_description = request.data.get('description', '')
        phase_steps = request.data.get('steps', [])

        project_type = project.analysis_scope.get('type', 'general') if project.analysis_scope else 'general'
        keywords = project.analysis_scope.get('keywords', []) if project.analysis_scope else []

        prompt = (
            f"You are an IP analytics expert helping with a {project_type} analysis.\n\n"
            f"Project: {project.name}\n"
            f"Description: {project.description or 'Not provided'}\n"
            f"Keywords: {', '.join(keywords) if keywords else 'Not specified'}\n\n"
            f"Current Phase: {phase_name} — {phase_description}\n\n"
            "Phase steps to complete:\n"
            + "\n".join(f"{i + 1}. {step}" for i, step in enumerate(phase_steps))
            + "\n\nBased on the project details, provide:\n"
            "1. 3-5 specific, actionable recommendations for completing this phase effectively\n"
            "2. 2-3 concrete next action items\n\n"
            'Respond ONLY with valid JSON in the format: {"recommendations": ["..."], "action_items": ["..."]}'
        )

        try:
            from ..models import LLMProviderConfig
            import json as json_lib

            # Try Anthropic first, then OpenAI, then fallback
            response_text = None
            for provider in ['anthropic', 'openai', 'google']:
                api_key = LLMProviderConfig.get_key(provider)
                if not api_key:
                    continue
                try:
                    if provider == 'anthropic':
                        import anthropic
                        client = anthropic.Anthropic(api_key=api_key)
                        msg = client.messages.create(
                            model='claude-haiku-4-5-20251001',
                            max_tokens=800,
                            messages=[{'role': 'user', 'content': prompt}],
                        )
                        response_text = msg.content[0].text
                    elif provider == 'openai':
                        import openai
                        client = openai.OpenAI(api_key=api_key)
                        resp = client.chat.completions.create(
                            model='gpt-4o-mini',
                            max_tokens=800,
                            messages=[{'role': 'user', 'content': prompt}],
                        )
                        response_text = resp.choices[0].message.content
                    break
                except Exception as e:
                    logger.warning(f"LLM provider {provider} failed: {e}")
                    continue

            if response_text:
                # Extract JSON from response
                start = response_text.find('{')
                end = response_text.rfind('}') + 1
                if start >= 0 and end > start:
                    data = json_lib.loads(response_text[start:end])
                    return Response({
                        'recommendations': data.get('recommendations', []),
                        'action_items': data.get('action_items', []),
                    })

            # Fallback: return contextual placeholder recommendations
            return Response({
                'recommendations': [
                    f'Review all available resources related to {phase_name} for this project type.',
                    'Consult relevant prior art databases and literature for this phase.',
                    'Document findings systematically to support downstream analysis phases.',
                ],
                'action_items': [
                    f'Begin with the first uncompleted step in the {phase_name} checklist.',
                    'Schedule a review session to validate outputs before moving to the next phase.',
                ],
            })

        except Exception as e:
            logger.error(f"Phase recommendations error: {e}")
            return Response(
                {'error': 'Failed to generate recommendations'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=['get', 'post'], url_path='workflow-progress')
    def workflow_progress(self, request, pk=None):
        """Get or update workflow progress for a project.

        GET: Returns current workflow progress.
        POST body (single step):
        {
            "phase_id": "data_collection",
            "step_index": 2,
            "completed": true
        }
        POST body (bulk):
        {
            "workflow_progress": { "phase_id": { "completed_steps": [0,1,2], "started_at": "...", "completed_at": "..." } }
        }
        """
        project = self.get_object()
        scope = project.analysis_scope or {}

        if request.method == 'GET':
            return Response({
                'project_id': str(project.id),
                'project_type': scope.get('type', ''),
                'workflow_progress': scope.get('workflow_progress', {}),
            })

        # POST — update workflow progress
        progress = scope.get('workflow_progress', {})

        # Bulk update
        if 'workflow_progress' in request.data:
            progress = request.data['workflow_progress']
        else:
            # Single step update
            phase_id = request.data.get('phase_id')
            step_index = request.data.get('step_index')
            completed = request.data.get('completed', True)

            if not phase_id or step_index is None:
                return Response({'error': 'phase_id and step_index required'}, status=status.HTTP_400_BAD_REQUEST)

            if phase_id not in progress:
                progress[phase_id] = {'completed_steps': [], 'started_at': timezone.now().isoformat()}

            steps = progress[phase_id].get('completed_steps', [])
            if completed and step_index not in steps:
                steps.append(step_index)
                steps.sort()
            elif not completed and step_index in steps:
                steps.remove(step_index)
            progress[phase_id]['completed_steps'] = steps

        scope['workflow_progress'] = progress
        project.analysis_scope = scope
        project.save(update_fields=['analysis_scope'])

        return Response({
            'project_id': str(project.id),
            'workflow_progress': progress,
        })

    @action(detail=True, methods=['post'])
    def start_analysis(self, request, pk=None):
        """Start automated analysis for a project"""
        project = self.get_object()
        
        # Update project status
        project.status = 'data_collection'
        project.save()
        
        # Initialize analytics data processor (would trigger background tasks)
        processor = AnalyticsDataProcessor(project)
        # processor.start_processing()  # Would start async processing
        
        return Response({'status': 'Analysis started', 'project_id': str(project.id)})
    
    @action(detail=True, methods=['get'])
    def export_data(self, request, pk=None):
        """Export project data in various formats"""
        project = self.get_object()
        format_type = request.query_params.get('format', 'json').lower()
        
        # Collect comprehensive project data
        project_data = {
            'project_info': {
                'id': str(project.id),
                'name': project.name,
                'description': project.description,
                'status': project.status,
                'priority': project.priority,
                'created_at': project.created_at.isoformat(),
                'updated_at': project.updated_at.isoformat(),
                'analysis_scope': project.analysis_scope,
                'progress_percentage': project.progress_percentage,
            }
        }
        
        # Add related data
        datasets = PatentDataset.objects.filter(project=project)
        visualizations = AnalyticsVisualization.objects.filter(project=project)
        reports = AnalyticsReport.objects.filter(project=project)
        insights = AnalyticsInsight.objects.filter(project=project)
        technology_areas = TechnologyArea.objects.filter(project=project)
        
        project_data.update({
            'datasets': [{
                'name': d.name,
                'description': d.description,
                'total_patents': d.total_patents,
                'processing_status': d.processing_status,
                'created_at': d.created_at.isoformat(),
            } for d in datasets],
            'visualizations': [{
                'title': v.title,
                'description': v.description,
                'visualization_type': v.visualization_type,
                'status': v.status,
                'created_at': v.created_at.isoformat(),
            } for v in visualizations],
            'reports': [{
                'title': r.title,
                'description': r.description,
                'status': r.status,
                'created_at': r.created_at.isoformat(),
            } for r in reports],
            'insights': [{
                'title': i.title,
                'description': i.description,
                'insight_type': i.insight_type,
                'confidence_score': i.confidence_score,
                'created_at': i.created_at.isoformat(),
            } for i in insights],
            'technology_areas': [{
                'name': ta.name,
                'description': ta.description,
                'patent_count': ta.patent_count,
                'keywords': ta.keywords,
            } for ta in technology_areas],
        })
        
        if format_type == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write project summary
            writer.writerow(['Project Export Summary'])
            writer.writerow(['Name', project.name])
            writer.writerow(['Description', project.description])
            writer.writerow(['Status', project.status])
            writer.writerow(['Priority', project.priority])
            writer.writerow(['Created', project.created_at.strftime('%Y-%m-%d')])
            writer.writerow([])
            
            # Write datasets
            writer.writerow(['Datasets'])
            writer.writerow(['Name', 'Description', 'Total Patents', 'Status'])
            for d in datasets:
                writer.writerow([d.name, d.description, d.total_patents, d.processing_status])
            writer.writerow([])
            
            # Write visualizations
            writer.writerow(['Visualizations'])
            writer.writerow(['Title', 'Type', 'Status', 'Created'])
            for v in visualizations:
                writer.writerow([v.title, v.visualization_type, v.status, v.created_at.strftime('%Y-%m-%d')])
            
            csv_content = output.getvalue()
            output.close()
            
            response = HttpResponse(csv_content, content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{project.name}_export.csv"'
            return response
            
        elif format_type == 'excel':
            try:
                import openpyxl
                from openpyxl import Workbook
                import io
                
                wb = Workbook()
                
                # Project Summary Sheet
                ws1 = wb.active
                ws1.title = "Project Summary"
                ws1['A1'] = 'Project Name'
                ws1['B1'] = project.name
                ws1['A2'] = 'Description'
                ws1['B2'] = project.description
                ws1['A3'] = 'Status'
                ws1['B3'] = project.status
                ws1['A4'] = 'Priority'
                ws1['B4'] = project.priority
                ws1['A5'] = 'Created'
                ws1['B5'] = project.created_at.strftime('%Y-%m-%d')
                
                # Datasets Sheet
                ws2 = wb.create_sheet("Datasets")
                ws2['A1'] = 'Name'
                ws2['B1'] = 'Description'
                ws2['C1'] = 'Total Patents'
                ws2['D1'] = 'Status'
                
                for idx, d in enumerate(datasets, start=2):
                    ws2[f'A{idx}'] = d.name
                    ws2[f'B{idx}'] = d.description
                    ws2[f'C{idx}'] = d.total_patents
                    ws2[f'D{idx}'] = d.processing_status
                
                # Visualizations Sheet
                ws3 = wb.create_sheet("Visualizations")
                ws3['A1'] = 'Title'
                ws3['B1'] = 'Type'
                ws3['C1'] = 'Status'
                ws3['D1'] = 'Created'
                
                for idx, v in enumerate(visualizations, start=2):
                    ws3[f'A{idx}'] = v.title
                    ws3[f'B{idx}'] = v.visualization_type
                    ws3[f'C{idx}'] = v.status
                    ws3[f'D{idx}'] = v.created_at.strftime('%Y-%m-%d')
                
                # Save to bytes
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)
                
                response = HttpResponse(
                    excel_file.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{project.name}_export.xlsx"'
                return response
                
            except ImportError:
                return Response({
                    'error': 'Excel export requires openpyxl package',
                    'message': 'Install openpyxl for Excel export functionality'
                }, status=400)
        
        # Default to JSON
        response = HttpResponse(
            json.dumps(project_data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{project.name}_export.json"'
        return response
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Create a duplicate of the project"""
        original_project = self.get_object()
        
        # Create new project with copied data
        new_project = AnalyticsProject.objects.create(
            name=f"{original_project.name} (Copy)",
            description=f"Copy of: {original_project.description}",
            status='draft',
            priority=original_project.priority,
            analysis_scope=original_project.analysis_scope,
            created_by=request.user if request.user.is_authenticated else original_project.created_by
        )
        
        # Copy related data
        # Technology Areas
        for tech_area in TechnologyArea.objects.filter(project=original_project):
            TechnologyArea.objects.create(
                project=new_project,
                name=tech_area.name,
                description=tech_area.description,
                keywords=tech_area.keywords,
                ipc_classes=tech_area.ipc_classes,
                cpc_classes=tech_area.cpc_classes,
                search_queries=tech_area.search_queries,
                confidence_threshold=tech_area.confidence_threshold,
            )
        
        # Copy visualizations (without chart_data to avoid large payloads)
        for viz in AnalyticsVisualization.objects.filter(project=original_project):
            AnalyticsVisualization.objects.create(
                project=new_project,
                title=f"{viz.title} (Copy)",
                description=viz.description,
                visualization_type=viz.visualization_type,
                config=viz.config,
                filters=viz.filters,
                width=viz.width,
                height=viz.height,
                is_interactive=viz.is_interactive,
                status='draft',
                created_by=request.user if request.user.is_authenticated else viz.created_by
            )
        
        return Response({
            'status': 'Project duplicated successfully',
            'new_project_id': str(new_project.id),
            'new_project_name': new_project.name
        })
    
    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        """Archive a project (set status to 'on_hold')"""
        project = self.get_object()
        previous_status = project.status
        
        project.status = 'on_hold'
        project.save()
        
        return Response({
            'status': 'Project archived successfully',
            'project_id': str(project.id),
            'previous_status': previous_status,
            'current_status': project.status
        })
    
    @action(detail=True, methods=['post'])
    def unarchive(self, request, pk=None):
        """Unarchive a project (restore previous status or set to 'active')"""
        project = self.get_object()
        previous_status = project.status
        
        # If the project was on hold, restore to active status
        if project.status == 'on_hold':
            # Get desired status from request, default to 'active'
            new_status = request.data.get('status', 'active')
            
            # Validate status is allowed
            valid_statuses = [choice[0] for choice in AnalyticsProject.STATUS_CHOICES]
            if new_status not in valid_statuses:
                return Response({
                    'error': f'Invalid status: {new_status}',
                    'valid_statuses': valid_statuses
                }, status=400)
            
            project.status = new_status
            project.save()
            
            return Response({
                'status': 'Project unarchived successfully',
                'project_id': str(project.id),
                'previous_status': previous_status,
                'current_status': project.status
            })
        else:
            return Response({
                'error': 'Project is not archived',
                'current_status': project.status
            }, status=400)
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update project status with workflow automation"""
        project = self.get_object()
        new_status = request.data.get('status')
        
        if not new_status:
            return Response({'error': 'Status is required'}, status=400)
        
        # Validate status is allowed
        valid_statuses = [choice[0] for choice in AnalyticsProject.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return Response({
                'error': f'Invalid status: {new_status}',
                'valid_statuses': valid_statuses
            }, status=400)
        
        previous_status = project.status
        project.status = new_status
        
        # Workflow automation: Set completed_date when status becomes 'completed'
        if new_status == 'completed' and previous_status != 'completed':
            from django.utils import timezone
            project.completed_date = timezone.now()
        
        project.save()
        
        return Response({
            'status': 'Project status updated successfully',
            'project_id': str(project.id),
            'previous_status': previous_status,
            'current_status': project.status,
            'completed_date': project.completed_date.isoformat() if project.completed_date else None
        })

    @action(detail=True, methods=['post'])
    def analyze_landscape(self, request, pk=None):
        """Dispatch landscape analysis to Celery and return task_id immediately."""
        from ..tasks import run_landscape_task
        project = self.get_object()
        task = run_landscape_task.delay(str(project.id))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'landscape_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_fto(self, request, pk=None):
        """Dispatch FTO analysis to Celery and return task_id immediately."""
        from ..tasks import run_fto_task
        project = self.get_object()
        target_description = request.data.get('target_description', '')
        task = run_fto_task.delay(str(project.id), target_description)
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'fto_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def find_white_space(self, request, pk=None):
        """Dispatch white-space analysis to Celery and return task_id immediately."""
        from ..tasks import run_white_space_task
        project = self.get_object()
        task = run_white_space_task.delay(str(project.id))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'white_space_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def forecast_trends(self, request, pk=None):
        """Dispatch trend analysis to Celery and return task_id immediately."""
        from ..tasks import run_trends_task
        project = self.get_object()
        task = run_trends_task.delay(str(project.id))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'trend_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def assess_portfolio(self, request, pk=None):
        """Dispatch portfolio assessment to Celery and return task_id immediately."""
        from ..tasks import run_portfolio_assessment_task
        project = self.get_object()
        task = run_portfolio_assessment_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'portfolio_assessment'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_market_analysis(self, request, pk=None):
        """Dispatch market analysis to Celery and return task_id immediately."""
        from ..tasks import run_market_analysis_task
        project = self.get_object()
        task = run_market_analysis_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'market_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_investment_analysis(self, request, pk=None):
        """Dispatch investment analysis to Celery and return task_id immediately.

        Optional POST params: income_value_per_patent, market_value_per_patent, cost_value_per_patent
        """
        from ..tasks import run_investment_analysis_task
        project = self.get_object()
        task = run_investment_analysis_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'investment_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_litigation_analysis(self, request, pk=None):
        """Dispatch litigation analysis to Celery and return task_id immediately."""
        from ..tasks import run_litigation_analysis_task
        project = self.get_object()
        task = run_litigation_analysis_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'litigation_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_licensing_analysis(self, request, pk=None):
        """Dispatch licensing analysis to Celery and return task_id immediately.

        Optional POST params: revenue_base_per_licensable, georgia_pacific_overrides
        """
        from ..tasks import run_licensing_analysis_task
        project = self.get_object()
        task = run_licensing_analysis_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'licensing_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['post'])
    def run_valuation_analysis(self, request, pk=None):
        """Dispatch valuation analysis to Celery and return task_id immediately.

        Optional POST params: revenue_base_per_patent, royalty_rate, discount_rate,
        reproduction_cost_per_patent, market_value_per_patent
        """
        from ..tasks import run_valuation_analysis_task
        project = self.get_object()
        task = run_valuation_analysis_task.delay(str(project.id), self._algo_kwargs(request))
        return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'valuation_analysis'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=False, methods=['get'], url_path='task-status/(?P<task_id>[^/.]+)')
    def task_status(self, request, task_id=None):
        """Poll Celery task status. Returns state + result when complete."""
        from celery.result import AsyncResult
        result = AsyncResult(task_id)
        if result.state == 'PENDING':
            return Response({'task_id': task_id, 'status': 'queued'})
        if result.state == 'STARTED':
            return Response({'task_id': task_id, 'status': 'running'})
        if result.state == 'SUCCESS':
            return Response({'task_id': task_id, 'status': 'completed', 'result': result.result})
        if result.state == 'FAILURE':
            return Response({'task_id': task_id, 'status': 'failed', 'error': str(result.result)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'task_id': task_id, 'status': result.state.lower()})

    @action(detail=True, methods=['post'], url_path='prosecution-analysis')
    def prosecution_analysis(self, request, pk=None):
        """Analyze prosecution history across project patents (M7 pipeline)."""
        from ..prosecution_service import analyze_project_prosecution

        project = self.get_object()
        max_patents = int(request.data.get('max_patents', 300))
        result = analyze_project_prosecution(str(project.id), max_patents=max_patents)
        self._persist_analysis(project, 'prosecution_analysis', result, request)
        return Response(result)

    @action(detail=True, methods=['post'])
    def family_analysis(self, request, pk=None):
        """Run patent family claim analysis for a project's patents.

        POST body:
        {
            "lens_id": "xxx-xxx-xxx",      # Lens ID of the patent to analyze
            "family_type": "simple",        # "simple" or "extended"
            "analysis_mode": "quick"        # "quick" (algorithmic) or "deep" (LLM)
        }
        """
        from ..family_analysis_service import FamilyAnalysisService
        from ..lens_service import LensAPIError

        lens_id = request.data.get('lens_id', '').strip()
        family_type = request.data.get('family_type', 'simple')
        analysis_mode = request.data.get('analysis_mode', 'quick')

        if not lens_id:
            return Response({'error': 'lens_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        project = self.get_object()

        try:
            service = FamilyAnalysisService()
            members = service.fetch_family_claims(lens_id, family_type)

            if len(members) < 2:
                return Response({
                    'error': 'Family has fewer than 2 members with data.',
                    'family_size': len(members),
                })

            if analysis_mode == 'deep':
                model = request.data.get('model', 'sonnet')
                result = service.deep_analysis(members, model)
            else:
                result = service.quick_analysis(members)

            # Persist the result
            PatentAnalysisResult.objects.update_or_create(
                application_id=lens_id,
                analysis_type='family_analysis',
                defaults={
                    'extracted_entities': result,
                    'metadata': {
                        'project_id': str(project.id),
                        'family_type': family_type,
                        'analysis_mode': analysis_mode,
                        'member_count': len(members),
                    },
                }
            )

            return Response(result)
        except LensAPIError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as exc:
            return Response({'error': f'Analysis failed: {str(exc)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='analysis-history')
    def analysis_history(self, request, pk=None):
        """Get past analysis results for a project."""
        project = self.get_object()
        analysis_type = request.query_params.get('type', '')

        qs = PatentAnalysisResult.objects.filter(
            application_id=str(project.id)
        ).order_by('-created_at')

        if analysis_type:
            qs = qs.filter(analysis_type=analysis_type)

        results = list(qs[:20].values(
            'id', 'analysis_type', 'metadata', 'created_at'
        ))

        return Response({'results': results})

    @action(detail=True, methods=['get'], url_path='analysis-history/(?P<result_id>[^/.]+)')
    def analysis_result_detail(self, request, pk=None, result_id=None):
        """Get a specific past analysis result."""
        try:
            result = PatentAnalysisResult.objects.get(id=result_id)
        except PatentAnalysisResult.DoesNotExist:
            return Response({'error': 'Result not found'}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'id': str(result.id),
            'analysis_type': result.analysis_type,
            'data': result.extracted_entities,
            'metadata': result.metadata,
            'created_at': result.created_at.isoformat(),
        })

    @action(detail=True, methods=['post'], url_path='generate-report')
    def generate_full_report(self, request, pk=None):
        """Generate a merged landscape + whitespace report with AI narrative.

        Runs both analyses, merges results, generates executive summary,
        and persists the report.
        """
        from ..narrative_service import generate_merged_report

        project = self.get_object()
        result = generate_merged_report(str(project.id))

        if 'error' in result.get('landscape', {}):
            return Response({'error': result['landscape']['error']}, status=status.HTTP_400_BAD_REQUEST)

        return Response(result)

    @action(detail=True, methods=['post'], url_path='fulltext-analysis')
    def fulltext_analysis(self, request, pk=None):
        """Run full text analysis on project patents (M6 pipeline)."""
        from ..fulltext_service import process_project_fulltext

        project = self.get_object()
        max_patents = int(request.data.get('max_patents', 200))
        result = process_project_fulltext(str(project.id), max_patents=max_patents)
        self._persist_analysis(project, 'fulltext_analysis', result, request)
        return Response(result)

    def _persist_analysis(self, project, analysis_type, result, request):
        """Persist an algorithm result for history, silently skipping on error."""
        if isinstance(result, dict) and 'error' in result:
            return
        try:
            PatentAnalysisResult.objects.create(
                application_id=str(project.id),
                analysis_type=analysis_type,
                extracted_entities=result,
                metadata={
                    'project_name': project.name,
                    'triggered_by': request.user.email if request.user.is_authenticated else 'system',
                },
            )
        except Exception as exc:
            logger.warning("Failed to persist %s result for project %s: %s", analysis_type, project.id, exc)

    @action(detail=True, methods=['post'])
    def run_bundle_analysis(self, request, pk=None):
        """Dispatch bundle analysis to Celery, or run synchronously if broker is unavailable."""
        from ..tasks import run_bundle_analysis_task
        from ..algorithms import run_bundle_analysis as _run_sync
        project = self.get_object()
        kwargs = self._bundle_algo_kwargs(request)
        try:
            task = run_bundle_analysis_task.delay(str(project.id), kwargs)
            # Persist the active task_id so the frontend can reconnect after a page reload
            ar = project.analysis_results or {}
            ar['bundle_analysis'] = {**(ar.get('bundle_analysis') or {}), 'task_id': task.id, 'task_status': 'queued'}
            project.analysis_results = ar
            project.save(update_fields=['analysis_results'])
            return Response({'task_id': task.id, 'status': 'queued', 'analysis_type': 'bundle_analysis'}, status=status.HTTP_202_ACCEPTED)
        except Exception:
            result = _run_sync(str(project.id), **(kwargs or {}))
            if 'error' in result:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)
            return Response(result)

    @action(detail=True, methods=['get'], url_path='bundle-analysis-result')
    def bundle_analysis_result(self, request, pk=None):
        """Return the last saved bundle analysis result and active task state for this project."""
        from celery.result import AsyncResult
        project = self.get_object()
        saved = (project.analysis_results or {}).get('bundle_analysis', {})
        if not saved:
            return Response({'has_result': False})

        task_id = saved.get('task_id')
        task_status = saved.get('task_status', 'unknown')

        # Refresh task status from Celery if we have a task_id
        if task_id:
            r = AsyncResult(task_id)
            if r.state == 'SUCCESS':
                task_status = 'completed'
            elif r.state == 'FAILURE':
                task_status = 'failed'
            elif r.state == 'PROGRESS':
                task_status = 'running'
                meta = r.info or {}
                return Response({
                    'has_result': 'assignment_matrix' in saved,
                    'task_id': task_id,
                    'task_status': task_status,
                    'progress': meta,
                    'result': saved if 'assignment_matrix' in saved else None,
                })
            elif r.state in ('PENDING', 'RECEIVED', 'STARTED'):
                task_status = 'running'

        return Response({
            'has_result': 'assignment_matrix' in saved,
            'task_id': task_id,
            'task_status': task_status,
            'result': saved if 'assignment_matrix' in saved else None,
        })

    @action(detail=True, methods=['post'])
    def extract_bundle_attributes(self, request, pk=None):
        """AI-extract bundle attributes for one or all patents in the project."""
        from ..tasks import run_bulk_attribute_extraction_task
        from ..bundle_attribute_service import extract_bundle_attributes_via_llm
        from ..models import PatentDataset, PatentRecord

        project = self.get_object()
        patent_record_ids = request.data.get('patent_record_ids')
        fields = request.data.get('fields')

        # Single patent → run synchronously
        if patent_record_ids and len(patent_record_ids) == 1:
            try:
                result = extract_bundle_attributes_via_llm(patent_record_ids[0], fields)
                return Response({'status': 'completed', 'extracted_count': len(result), 'result': result})
            except ValueError as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Bulk → Celery
        task = run_bulk_attribute_extraction_task.delay(str(project.id), patent_record_ids, fields)
        return Response({'task_id': task.id, 'status': 'queued'}, status=status.HTTP_202_ACCEPTED)

    @action(detail=True, methods=['get', 'patch'])
    def bundle_attributes(self, request, pk=None):
        """GET all PatentBundleAttributes for project patents. PATCH to update a patent's attributes."""
        from ..models import PatentBundleAttributes, PatentDataset, PatentRecord
        from ..serializers import PatentBundleAttributesSerializer
        from ..bundle_attribute_service import update_attributes_manually

        project = self.get_object()

        if request.method == 'GET':
            # Gather all patent_record ids for this project's datasets
            dataset_ids = PatentDataset.objects.filter(project=project).values_list('id', flat=True)
            record_ids = PatentRecord.objects.filter(dataset_id__in=dataset_ids).values_list('id', flat=True)
            qs = PatentBundleAttributes.objects.filter(patent_record_id__in=record_ids).select_related('patent_record')
            total = qs.count()
            limit = min(int(request.query_params.get('limit', 50)), 200)
            offset = int(request.query_params.get('offset', 0))
            serializer = PatentBundleAttributesSerializer(qs[offset:offset + limit], many=True)
            return Response({'count': total, 'results': serializer.data})

        # PATCH
        patent_record_id = request.data.get('patent_record_id')
        attributes = request.data.get('attributes', {})
        if not patent_record_id:
            return Response({'error': 'patent_record_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            updated = update_attributes_manually(patent_record_id, attributes)
            return Response({'updated': updated})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='from-portfolio')
    def from_portfolio(self, request):
        """Get or create an analytics project linked to a portfolio, returning the project id."""
        from domains.patents.models import Portfolio
        portfolio_id = request.data.get('portfolio_id')
        if not portfolio_id:
            return Response({'error': 'portfolio_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            portfolio = Portfolio.objects.get(pk=portfolio_id)
        except Portfolio.DoesNotExist:
            return Response({'error': 'Portfolio not found'}, status=status.HTTP_404_NOT_FOUND)

        # Find an existing project linked to this portfolio owned by this user
        existing = AnalyticsProject.objects.filter(
            portfolio=portfolio, created_by=request.user
        ).order_by('-created_at').first()
        if existing:
            return Response({'project_id': str(existing.id), 'created': False, 'project_name': existing.name})

        # Create a lightweight project linked to the portfolio
        project = AnalyticsProject.objects.create(
            name=f'{portfolio.name} — Bundle Analysis',
            description='Auto-created for bundle analysis from portfolio.',
            status='active',
            portfolio=portfolio,
            created_by=request.user,
        )
        return Response({'project_id': str(project.id), 'created': True, 'project_name': project.name}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='from-excel-upload')
    def from_excel_upload(self, request):
        """Create an analytics project from an uploaded Excel/CSV file.

        Multipart POST fields:
          file  — the spreadsheet file (.xlsx, .xls, .csv)
          name  — optional friendly name for the project/dataset
        """
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        ext = uploaded_file.name.rsplit('.', 1)[-1].lower() if '.' in uploaded_file.name else ''
        if ext not in ('xlsx', 'xls', 'csv'):
            return Response({'error': 'Only .xlsx, .xls, and .csv files are supported'}, status=status.HTTP_400_BAD_REQUEST)

        friendly_name = (request.data.get('name') or '').strip()
        base_name = friendly_name or uploaded_file.name.rsplit('.', 1)[0]
        project_name = f'{base_name} — Bundle Analysis'

        # Create a lightweight project (no portfolio linkage)
        project = AnalyticsProject.objects.create(
            name=project_name,
            description=f'Auto-created from uploaded file: {uploaded_file.name}',
            status='active',
            created_by=request.user,
        )

        # Create the dataset record
        dataset = PatentDataset.objects.create(
            project=project,
            name=base_name,
            data_source='manual_upload',
            data_file=uploaded_file,
            processing_status='pending',
        )

        # Process synchronously (same as datasets process_data action)
        proc_result = process_patent_dataset(str(dataset.id))
        dataset.refresh_from_db()

        return Response({
            'project_id': str(project.id),
            'project_name': project.name,
            'dataset_id': str(dataset.id),
            'processing_status': dataset.processing_status,
            'total_patents': dataset.total_patents,
            'created': True,
        }, status=status.HTTP_201_CREATED)

    def _bundle_algo_kwargs(self, request):
        if not request.data:
            return {}
        safe_keys = {'config_id', 'preset', 'thresholds', 'enabled_bundles', 'persist_assignments'}
        return {k: v for k, v in request.data.items() if k in safe_keys}

    def _algo_kwargs(self, request):
        """Extract algorithm-safe kwargs from request data.

        Only passes through known numeric/dict/list params to prevent injection.
        """
        if not request.data:
            return {}
        safe_keys = {
            'income_value_per_patent', 'market_value_per_patent', 'cost_value_per_patent',
            'revenue_base_per_patent', 'royalty_rate', 'discount_rate',
            'reproduction_cost_per_patent', 'revenue_base_per_licensable',
            'georgia_pacific_overrides', 'target_description',
        }
        return {k: v for k, v in request.data.items() if k in safe_keys}

    # ── Watch / Automation endpoints ──────────────────────────────────

    @action(detail=True, methods=['post'], url_path='create-watch')
    def create_watch(self, request, pk=None):
        """Save a research query as a watch for periodic re-run."""
        from ..models import ResearchQuery

        project = self.get_object()
        query_id = request.data.get('query_id')
        cadence = request.data.get('cadence', 'weekly')
        thresholds = request.data.get('thresholds', {'new_patents_min': 5, 'new_assignee': True})

        if not query_id:
            return Response({'error': 'query_id required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            query = ResearchQuery.objects.get(id=query_id, project=project)
        except ResearchQuery.DoesNotExist:
            return Response({'error': 'Query not found'}, status=status.HTTP_404_NOT_FOUND)

        query.is_watch = True
        query.watch_cadence = cadence
        query.alert_thresholds = thresholds
        query.save(update_fields=['is_watch', 'watch_cadence', 'alert_thresholds'])

        return Response({
            'query_id': str(query.id),
            'is_watch': True,
            'cadence': cadence,
            'thresholds': thresholds,
        })

    @action(detail=True, methods=['get'], url_path='watches')
    def list_watches(self, request, pk=None):
        """List all active watches for a project."""
        from ..models import ResearchQuery

        project = self.get_object()
        watches = ResearchQuery.objects.filter(
            project=project, is_watch=True
        ).values(
            'id', 'query_name', 'watch_cadence', 'alert_thresholds',
            'last_watch_run', 'watch_diff', 'created_at',
        )

        results = list(watches)
        for w in results:
            w['id'] = str(w['id'])

        return Response({'watches': results})

    @action(detail=True, methods=['post'], url_path='run-watch')
    def run_watch(self, request, pk=None):
        """Manually trigger a watch re-run for a specific query."""
        from ..models import ResearchQuery
        from ..watch_service import run_watch as execute_watch

        project = self.get_object()
        query_id = request.data.get('query_id')

        if not query_id:
            return Response({'error': 'query_id required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            query = ResearchQuery.objects.get(id=query_id, project=project, is_watch=True)
        except ResearchQuery.DoesNotExist:
            return Response({'error': 'Watch query not found'}, status=status.HTTP_404_NOT_FOUND)

        result = execute_watch(query)
        return Response(result)

    @action(detail=True, methods=['post'], url_path='delete-watch')
    def delete_watch(self, request, pk=None):
        """Remove a query from the watch list."""
        from ..models import ResearchQuery

        project = self.get_object()
        query_id = request.data.get('query_id')

        if not query_id:
            return Response({'error': 'query_id required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            query = ResearchQuery.objects.get(id=query_id, project=project)
        except ResearchQuery.DoesNotExist:
            return Response({'error': 'Query not found'}, status=status.HTTP_404_NOT_FOUND)

        query.is_watch = False
        query.watch_cadence = ''
        query.save(update_fields=['is_watch', 'watch_cadence'])

        return Response({'query_id': str(query.id), 'is_watch': False})

    # ─────────────────────────────────────────────────────────────────────────
    # Sales Package endpoints
    # ─────────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=['get', 'post'], url_path='sales-packages')
    def sales_packages(self, request, pk=None):
        """List or create SalesPackages for a project."""
        project = self.get_object()

        if request.method == 'GET':
            qs = SalesPackage.objects.filter(project=project)
            serializer = SalesPackageSerializer(qs, many=True)
            return Response(serializer.data)

        # POST — create
        serializer = SalesPackageSerializer(data={**request.data, 'project': str(project.id)})
        if serializer.is_valid():
            serializer.save(project=project, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['patch', 'delete'],
            url_path=r'sales-packages/(?P<package_id>[^/.]+)')
    def sales_package_detail(self, request, pk=None, package_id=None):
        """Update or delete a specific SalesPackage."""
        project = self.get_object()
        pkg = get_object_or_404(SalesPackage, id=package_id, project=project)

        if request.method == 'DELETE':
            pkg.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        serializer = SalesPackageSerializer(pkg, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='export-sales-package')
    def export_sales_package(self, request, pk=None):
        """Export a sales package as Excel, JSON, or PDF."""
        import io
        from ..patent_data_service import get_project_patents
        from ..models import BundleQualityScore, BundleType, PatentBundleAttributes

        project = self.get_object()
        fmt = request.data.get('format', 'excel')
        bundle_codes = request.data.get('bundle_codes', [])
        package_name = request.data.get('package_name', 'Sales Package')

        # Load analysis result
        analysis = (project.analysis_results or {}).get('bundle_analysis', {})
        pat_summary = {
            p['patent_record_id']: p
            for p in analysis.get('patent_attribute_summary', [])
        }
        scorecard_rows = {
            s['bundle_code']: s
            for s in analysis.get('quality_scorecard', [])
        }

        # Filter to selected bundles
        if bundle_codes:
            scorecard_rows = {k: v for k, v in scorecard_rows.items() if k in bundle_codes}

        # Collect patents in selected bundles
        assignment_matrix = analysis.get('assignment_matrix', {})
        matrix = assignment_matrix.get('matrix', [])
        patent_ids_in_matrix = assignment_matrix.get('patent_ids', [])
        bundle_codes_in_matrix = assignment_matrix.get('bundle_codes', [])

        selected_patent_ids = set()
        for bundle_code in (bundle_codes or bundle_codes_in_matrix):
            if bundle_code in bundle_codes_in_matrix:
                col_idx = bundle_codes_in_matrix.index(bundle_code)
                for row_idx, row in enumerate(matrix):
                    if row[col_idx]:
                        selected_patent_ids.add(patent_ids_in_matrix[row_idx])

        # Build patent rows from PatentBundleAttributes + PatentRecord
        patent_rows = []
        record_qs = PatentRecord.objects.filter(id__in=selected_patent_ids).select_related() if selected_patent_ids else PatentRecord.objects.none()
        attr_map = {
            str(a.patent_record_id): a
            for a in PatentBundleAttributes.objects.filter(patent_record_id__in=selected_patent_ids)
        } if selected_patent_ids else {}

        for pr in record_qs:
            a = attr_map.get(str(pr.id))
            summary = pat_summary.get(str(pr.id), {})
            patent_rows.append({
                'patent_id': pr.patent_id or str(pr.id),
                'title': pr.title or '',
                'assignee': pr.assignee or '',
                'filing_date': str(pr.filing_date or ''),
                'grant_date': str(pr.grant_date or ''),
                'legal_status': pr.legal_status or '',
                'ipc': pr.ipc_classification or '',
                'bundle_codes': ', '.join(summary.get('bundle_codes', [])),
                'h1_claim_strength': getattr(a, 'h1_claim_strength', '') if a else '',
                'h2_prior_art_exposure': getattr(a, 'h2_prior_art_exposure', '') if a else '',
                'h9_eou_availability': getattr(a, 'h9_eou_availability', '') if a else '',
                'e4_remaining_term_years': getattr(a, 'e4_remaining_term_years', '') if a else '',
                'd1_external_detectability': getattr(a, 'd1_external_detectability', '') if a else '',
                'c2_breadth': getattr(a, 'c2_breadth', '') if a else '',
                'b1_sep_potential': getattr(a, 'b1_sep_potential', '') if a else '',
                'h5_forward_citations': getattr(a, 'h5_forward_citations', '') if a else '',
                'attribute_source': summary.get('attribute_source', ''),
                'pct_filled': summary.get('pct_filled', 0),
            })

        if fmt == 'json':
            payload = {
                'package_name': package_name,
                'project_id': str(project.id),
                'bundle_codes': bundle_codes,
                'total_patents': len(patent_rows),
                'quality_scorecard': list(scorecard_rows.values()),
                'patents': patent_rows,
            }
            resp = HttpResponse(
                json.dumps(payload, indent=2, default=str),
                content_type='application/json'
            )
            resp['Content-Disposition'] = f'attachment; filename="{package_name}.json"'
            return resp

        if fmt == 'pdf':
            try:
                from utils.export import PDFExporter
                pdf = PDFExporter(title=package_name)
                # Summary section
                pdf.add_section('Package Summary', [
                    f'Project: {project.name}',
                    f'Total patents: {len(patent_rows)}',
                    f'Bundles included: {", ".join(bundle_codes) or "All"}',
                ])
                # Scorecard table
                scorecard_data = [[
                    s.get('bundle_name', s.get('bundle_code', '')),
                    str(s.get('patent_count', 0)),
                    s.get('strength_flag', ''),
                    f"{s.get('avg_claim_strength') or 0:.1f}",
                    f"{s.get('avg_remaining_term') or 0:.1f}y",
                ] for s in scorecard_rows.values()]
                pdf.add_table(
                    ['Bundle', 'Patents', 'Quality', 'Avg Strength', 'Avg Term'],
                    scorecard_data,
                    title='Bundle Scorecard'
                )
                resp = pdf.get_response(filename=f'{package_name}.pdf')
                return resp
            except Exception as e:
                return Response({'error': f'PDF generation failed: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Default: Excel
        try:
            from utils.export import ExcelExporter
            exporter = ExcelExporter()

            # Sheet 1: Summary
            summary_data = [
                ['Package Name', package_name],
                ['Project', project.name],
                ['Total Patents', len(patent_rows)],
                ['Bundles Included', len(bundle_codes) if bundle_codes else 'All'],
                ['Exported At', timezone.now().strftime('%Y-%m-%d %H:%M UTC')],
            ]
            exporter.add_sheet('Summary').write_data(['Field', 'Value'], summary_data)

            # Sheet 2: Patent List
            patent_headers = [
                'Patent ID', 'Title', 'Assignee', 'Filing Date', 'Grant Date',
                'Legal Status', 'IPC', 'Bundles',
                'Claim Strength (H1)', 'Prior Art Risk (H2)', 'EoU Availability (H9)',
                'Remaining Term (E4, yrs)', 'Detectability (D1)', 'Breadth (C2)',
                'SEP Potential (B1)', 'Forward Citations (H5)',
                'Attribute Source', '% Filled',
            ]
            patent_data = [[
                r['patent_id'], r['title'], r['assignee'], r['filing_date'], r['grant_date'],
                r['legal_status'], r['ipc'], r['bundle_codes'],
                r['h1_claim_strength'], r['h2_prior_art_exposure'], r['h9_eou_availability'],
                r['e4_remaining_term_years'], r['d1_external_detectability'], r['c2_breadth'],
                r['b1_sep_potential'], r['h5_forward_citations'],
                r['attribute_source'], r['pct_filled'],
            ] for r in patent_rows]
            exporter.add_sheet('Patent List').write_data(patent_headers, patent_data)

            # Sheet 3: Bundle Scorecard
            score_headers = [
                'Bundle Code', 'Bundle Name', 'Patent Count', 'Quality',
                'Avg Claim Strength', 'Avg Breadth', 'Trilateral %', 'Avg Term (yrs)',
                'Avg Detectability', 'Avg Forward Citations', 'SEP %', 'Continuation %',
                'Weakest H1', 'Invalidity Exposure %', 'EoU Ready %', 'Survived %', 'Cont. Optionality %',
            ]
            score_data = [[
                s.get('bundle_code', ''), s.get('bundle_name', ''),
                s.get('patent_count', 0), s.get('strength_flag', ''),
                s.get('avg_claim_strength', ''), s.get('avg_breadth', ''),
                s.get('pct_trilateral', ''), s.get('avg_remaining_term', ''),
                s.get('avg_detectability', ''), s.get('avg_forward_citations', ''),
                s.get('pct_sep', ''), s.get('pct_continuation_live', ''),
                s.get('gate_weakest_h1', ''), s.get('gate_invalidity_exposure_pct', ''),
                s.get('gate_eou_ready_pct', ''), s.get('gate_survived_pct', ''),
                s.get('gate_cont_optionality_pct', ''),
            ] for s in scorecard_rows.values()]
            exporter.add_sheet('Bundle Scorecard').write_data(score_headers, score_data)

            # Sheet 4: Attribute Data (full 42 fields for selected patents)
            attr_headers = [
                'Patent ID',
                'A1 Domain', 'A2 Subcategory', 'A3 Stack Layer', 'A4 Subsystem', 'A5 Use Case',
                'B1 SEP Potential', 'B2 Standard Tagged', 'B3 Interface Role',
                'C1 Claim Type', 'C2 Breadth', 'C3 Claim Count', 'C4 Design-Around Difficulty',
                'D1 External Detectability', 'D2 Teardown Detectability', 'D3 Reads on Products',
                'E1 Family Size', 'E2 Prosecution Status', 'E3 Continuation', 'E4 Remaining Term', 'E5 Maintenance',
                'F1 Jurisdictions', 'F2 Trilateral', 'F3 Major Market Score',
                'G1 Convergence Theme', 'G2 Generation Tag', 'G3 Cross-Industry',
                'H1 Claim Strength', 'H2 Prior Art Exposure', 'H3 Prosecution Risk',
                'H4 Divided Infringement', 'H5 Forward Citations', 'H6 Backward Citations',
                'H7 Litigation History', 'H8 Chain of Title', 'H9 EoU Availability', 'H10 Encumbrance',
                'I1 Product Mapping', 'I2 Implementation Maturity', 'I3 Adjacent Market', 'I4 Workaround Complexity',
                'AI Extracted Fields', 'Manual Fields', 'Derived Fields',
            ]
            attr_data = []
            for pr in record_qs:
                a = attr_map.get(str(pr.id))
                attr_data.append([
                    pr.patent_id or str(pr.id),
                    getattr(a, 'a1_primary_domain', '') if a else '',
                    getattr(a, 'a2_tech_subcategory', '') if a else '',
                    getattr(a, 'a3_stack_layer', '') if a else '',
                    getattr(a, 'a4_subsystem', '') if a else '',
                    getattr(a, 'a5_use_case', '') if a else '',
                    getattr(a, 'b1_sep_potential', '') if a else '',
                    getattr(a, 'b2_standard_tagged', '') if a else '',
                    getattr(a, 'b3_interface_role', '') if a else '',
                    getattr(a, 'c1_claim_type', '') if a else '',
                    getattr(a, 'c2_breadth', '') if a else '',
                    getattr(a, 'c3_claim_count', '') if a else '',
                    getattr(a, 'c4_design_around_difficulty', '') if a else '',
                    getattr(a, 'd1_external_detectability', '') if a else '',
                    getattr(a, 'd2_teardown_detectability', '') if a else '',
                    getattr(a, 'd3_reads_on_products', '') if a else '',
                    getattr(a, 'e1_family_size', '') if a else '',
                    getattr(a, 'e2_prosecution_status', '') if a else '',
                    getattr(a, 'e3_continuation', '') if a else '',
                    getattr(a, 'e4_remaining_term_years', '') if a else '',
                    getattr(a, 'e5_maintenance_status', '') if a else '',
                    ', '.join(getattr(a, 'f1_jurisdictions', None) or []) if a else '',
                    getattr(a, 'f2_trilateral', '') if a else '',
                    getattr(a, 'f3_major_market_score', '') if a else '',
                    getattr(a, 'g1_convergence_theme', '') if a else '',
                    getattr(a, 'g2_generation_tag', '') if a else '',
                    getattr(a, 'g3_cross_industry_applicability', '') if a else '',
                    getattr(a, 'h1_claim_strength', '') if a else '',
                    getattr(a, 'h2_prior_art_exposure', '') if a else '',
                    getattr(a, 'h3_prosecution_risk', '') if a else '',
                    getattr(a, 'h4_divided_infringement_risk', '') if a else '',
                    getattr(a, 'h5_forward_citations', '') if a else '',
                    getattr(a, 'h6_backward_citations', '') if a else '',
                    getattr(a, 'h7_litigation_history', '') if a else '',
                    getattr(a, 'h8_chain_of_title', '') if a else '',
                    getattr(a, 'h9_eou_availability', '') if a else '',
                    getattr(a, 'h10_encumbrance_status', '') if a else '',
                    getattr(a, 'i1_product_mapping_confidence', '') if a else '',
                    getattr(a, 'i2_implementation_maturity', '') if a else '',
                    getattr(a, 'i3_adjacent_market_reread', '') if a else '',
                    getattr(a, 'i4_workaround_complexity', '') if a else '',
                    ', '.join(getattr(a, 'ai_extracted_fields', None) or []) if a else '',
                    ', '.join(getattr(a, 'manually_set_fields', None) or []) if a else '',
                    ', '.join(getattr(a, 'derived_fields', None) or []) if a else '',
                ])
            exporter.add_sheet('Attribute Data').write_data(attr_headers, attr_data)

            return exporter.get_response(filename=f'{package_name}.xlsx')

        except Exception as e:
            logger.exception('Sales package export failed')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='generate-listing')
    def generate_listing(self, request, pk=None):
        """Generate a Teaser and Value Proposition listing for a SalesPackage.

        Body: { package_id: str, pattern_override: 'A'|'B'|'C'|'D'|null }
        """
        from ..models import PatentBundleAttributes
        from ..narrative_service import generate_value_proposition

        project = self.get_object()
        package_id = request.data.get('package_id')
        pattern_override = request.data.get('pattern_override') or None

        if not package_id:
            return Response({'error': 'package_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        pkg = get_object_or_404(SalesPackage, id=package_id, project=project)

        # Apply pattern override
        if pattern_override and pattern_override in ('A', 'B', 'C', 'D'):
            pkg.listing_pattern = pattern_override
            pkg.save(update_fields=['listing_pattern'])

        # Load bundle analysis result from cached project analysis
        analysis = (project.analysis_results or {}).get('bundle_analysis', {})
        assignment_matrix = analysis.get('assignment_matrix', {})
        matrix = assignment_matrix.get('matrix', [])
        patent_ids_in_matrix = assignment_matrix.get('patent_ids', [])
        bundle_codes_in_matrix = assignment_matrix.get('bundle_codes', [])

        # Collect patent IDs in selected bundles
        selected_patent_ids = set()
        for code in (pkg.bundle_codes or []):
            if code in bundle_codes_in_matrix:
                col_idx = bundle_codes_in_matrix.index(code)
                for row_idx, row in enumerate(matrix):
                    if row[col_idx]:
                        selected_patent_ids.add(patent_ids_in_matrix[row_idx])

        # Load top patent attributes (cap at 20 for prompt size)
        attributes_qs = PatentBundleAttributes.objects.filter(
            project=project,
            patent_record_id__in=list(selected_patent_ids)[:20],
        ).values() if selected_patent_ids else PatentBundleAttributes.objects.filter(
            project=project
        ).values()[:20]

        patent_attributes = list(attributes_qs)

        bundle_data = {
            'assignments': assignment_matrix,
            'scorecard': {
                row['bundle_code']: row
                for row in analysis.get('quality_scorecard', [])
            },
            'composition': analysis.get('bundle_composition', {}),
        }

        # Generate
        try:
            result = generate_value_proposition(pkg, bundle_data, patent_attributes)
        except Exception as e:
            logger.exception('generate_value_proposition failed')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Persist to package
        pkg.generated_teaser = result['teaser']
        pkg.generated_listing = result['listing']
        pkg.listing_tier_report = result['tier_report']
        pkg.listing_generated_at = timezone.now()
        if not pkg.listing_pattern:
            pkg.listing_pattern = result['pattern_used']
        pkg.meta_tags = result.get('meta_tags')
        pkg.lint_results = result.get('lint_results')
        pkg.quality_gates = result.get('quality_gates')
        pkg.tier_validation = result.get('tier_validation')
        pkg.suggested_archetype = result.get('suggested_archetype', '')
        pkg.archetype_reason = result.get('archetype_reason', '')
        pkg.save(update_fields=[
            'generated_teaser', 'generated_listing',
            'listing_tier_report', 'listing_generated_at', 'listing_pattern',
            'meta_tags', 'lint_results', 'quality_gates', 'tier_validation',
            'suggested_archetype', 'archetype_reason',
        ])

        return Response({
            'teaser': pkg.generated_teaser,
            'listing': pkg.generated_listing,
            'tier_report': pkg.listing_tier_report,
            'suggested_pattern': result['suggested_pattern'],
            'pattern_used': result['pattern_used'],
            'meta_tags': pkg.meta_tags,
            'lint_results': pkg.lint_results,
            'quality_gates': pkg.quality_gates,
            'tier_validation': pkg.tier_validation,
            'suggested_archetype': pkg.suggested_archetype,
            'archetype_reason': pkg.archetype_reason,
        })

    @action(detail=True, methods=['post'], url_path='generate-deck')
    def generate_deck(self, request, pk=None):
        """Generate Rung 3 non-confidential offering deck for a SalesPackage.

        Body: { package_id: str }
        """
        from ..models import PatentBundleAttributes
        from ..narrative_service import generate_deck as _generate_deck

        project = self.get_object()
        package_id = request.data.get('package_id')
        if not package_id:
            return Response({'error': 'package_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        pkg = get_object_or_404(SalesPackage, id=package_id, project=project)

        analysis = (project.analysis_results or {}).get('bundle_analysis', {})
        assignment_matrix = analysis.get('assignment_matrix', {})
        matrix = assignment_matrix.get('matrix', [])
        patent_ids_in_matrix = assignment_matrix.get('patent_ids', [])
        bundle_codes_in_matrix = assignment_matrix.get('bundle_codes', [])

        selected_patent_ids = set()
        for code in (pkg.bundle_codes or []):
            if code in bundle_codes_in_matrix:
                col_idx = bundle_codes_in_matrix.index(code)
                for row_idx, row in enumerate(matrix):
                    if row[col_idx]:
                        selected_patent_ids.add(patent_ids_in_matrix[row_idx])

        attributes_qs = PatentBundleAttributes.objects.filter(
            project=project,
            patent_record_id__in=list(selected_patent_ids)[:20],
        ).values() if selected_patent_ids else PatentBundleAttributes.objects.filter(
            project=project
        ).values()[:20]

        patent_attributes = list(attributes_qs)
        bundle_data = {
            'assignments': assignment_matrix,
            'scorecard': {row['bundle_code']: row for row in analysis.get('quality_scorecard', [])},
            'composition': analysis.get('bundle_composition', {}),
        }

        try:
            deck_md = _generate_deck(pkg, bundle_data, patent_attributes, pkg.generated_listing or '')
        except Exception as e:
            logger.exception('generate_deck failed')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        pkg.generated_deck = deck_md
        pkg.save(update_fields=['generated_deck'])
        return Response({'deck': deck_md})

    @action(detail=True, methods=['post'], url_path='generate-cim')
    def generate_cim(self, request, pk=None):
        """Generate Rung 4 CIM outline for a SalesPackage.

        Body: { package_id: str }
        """
        from ..models import PatentBundleAttributes
        from ..narrative_service import generate_cim as _generate_cim

        project = self.get_object()
        package_id = request.data.get('package_id')
        if not package_id:
            return Response({'error': 'package_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        pkg = get_object_or_404(SalesPackage, id=package_id, project=project)

        analysis = (project.analysis_results or {}).get('bundle_analysis', {})
        assignment_matrix = analysis.get('assignment_matrix', {})
        matrix = assignment_matrix.get('matrix', [])
        patent_ids_in_matrix = assignment_matrix.get('patent_ids', [])
        bundle_codes_in_matrix = assignment_matrix.get('bundle_codes', [])

        selected_patent_ids = set()
        for code in (pkg.bundle_codes or []):
            if code in bundle_codes_in_matrix:
                col_idx = bundle_codes_in_matrix.index(code)
                for row_idx, row in enumerate(matrix):
                    if row[col_idx]:
                        selected_patent_ids.add(patent_ids_in_matrix[row_idx])

        attributes_qs = PatentBundleAttributes.objects.filter(
            project=project,
            patent_record_id__in=list(selected_patent_ids)[:20],
        ).values() if selected_patent_ids else PatentBundleAttributes.objects.filter(
            project=project
        ).values()[:20]

        patent_attributes = list(attributes_qs)
        bundle_data = {
            'assignments': assignment_matrix,
            'scorecard': {row['bundle_code']: row for row in analysis.get('quality_scorecard', [])},
            'composition': analysis.get('bundle_composition', {}),
        }

        try:
            cim_md = _generate_cim(pkg, bundle_data, patent_attributes)
        except Exception as e:
            logger.exception('generate_cim failed')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        pkg.generated_cim = cim_md
        pkg.save(update_fields=['generated_cim'])
        return Response({'cim': cim_md})

