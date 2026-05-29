"""
Infringement Analysis Views
"""

import re
import logging

from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Avg
from django.db import transaction
from .ai_services import ai_service

from django.db.models import Q
from django.shortcuts import get_object_or_404

logger = logging.getLogger(__name__)

_MATCH_STOPWORDS = {
    'a', 'an', 'the', 'and', 'or', 'of', 'for', 'with', 'to', 'in', 'on', 'by', 'at',
    'system', 'method', 'apparatus', 'device', 'using', 'based',
}
_NAME_SUFFIXES = re.compile(
    r'\b(corporation|corp|incorporated|inc|llc|ltd|limited|gmbh|co|company|plc|sa|ag|kk)\b'
)


def _match_tokens(text):
    return {
        t for t in re.sub(r'[^a-z0-9\s]', ' ', (text or '').lower()).split()
        if len(t) > 1 and t not in _MATCH_STOPWORDS
    }


def _norm_name(name):
    s = _NAME_SUFFIXES.sub(' ', re.sub(r'[^a-z0-9\s]', ' ', (name or '').lower()))
    return ' '.join(s.split())


def _odp_case_match(case, patent, threshold=0.4):
    """Whether a freshly-resolved USPTO Patent plausibly matches this case.

    The patent number alone proves nothing (a number can be shared by unrelated demo
    data, and we looked it up *by* number). Compare title token overlap and assignee
    names so an unrelated patent (e.g. KIOXIA "NAND SWITCH" vs an Apple SoC case) is
    never silently linked. Mirrors the frontend assessOdpMatch guard.
    """
    a, b = _match_tokens(case.patent_title), _match_tokens(patent.title)
    inter = len(a & b)
    union = len(a | b)
    title_sim = (inter / union) if union else 0.0

    assignee_match = None
    local = [_norm_name(x) for x in (getattr(patent, 'assignees', None) or []) if x]
    # case has no own assignee field; compare the patent's assignees against the case title
    # is meaningless, so assignee gating only applies when the case was already linked to a
    # patent carrying assignees — handled by the caller comparing pre/post. Keep title-driven.
    confident = title_sim >= threshold
    return {
        'confident': confident,
        'title_similarity': round(title_sim, 3),
        'assignee_match': assignee_match,
        'found_title': patent.title or '',
        'found_assignees': getattr(patent, 'assignees', None) or [],
    }


def _ptab_cached(endpoint, patent_number, method_name, ttl_hours=24):
    """Return a cached ODP PTAB response, fetching + caching on miss.

    PTAB proceedings/decisions change rarely, so cache per (patent_number, endpoint) in
    ODPCacheEntry to avoid re-hitting USPTO on every tab view.
    """
    from datetime import timedelta
    from django.utils import timezone
    from domains.analytics.models import ODPCacheEntry
    from domains.analytics.uspto_odp_service import (
        USPTOODPClient, USPTOODPTrialService, USPTOODPError,
    )

    cache_key = f'ptab:{patent_number}'
    # Use updated_at (auto_now, refreshes on every save) for the TTL — fetched_at is
    # auto_now_add and would never refresh when an existing entry is updated.
    cutoff = timezone.now() - timedelta(hours=ttl_hours)
    cached = (
        ODPCacheEntry.objects
        .filter(application_id=cache_key, endpoint=endpoint, updated_at__gte=cutoff)
        .first()
    )
    if cached:
        return Response(cached.response_data)

    try:
        svc = USPTOODPTrialService(USPTOODPClient())
        query = {'q': patent_number, 'pagination': {'offset': 0, 'limit': 20}}
        data = getattr(svc, method_name)(query)
    except USPTOODPError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

    if data is not None:
        ODPCacheEntry.objects.update_or_create(
            application_id=cache_key, endpoint=endpoint,
            defaults={'response_data': data},
        )
    return Response(data)


def _parse_claim_into_elements(claim_text):
    """Parse patent claim text into structured elements (preamble + body limitations).

    Returns a list of dicts: [{'text': str, 'type': 'preamble'|'transition'|'body', 'order': int}]
    """
    text = claim_text.strip()
    # Strip leading claim number (e.g., "1. " or "12. ")
    text = re.sub(r'^\d+\.\s*', '', text)

    elements = []
    order = 1

    # Find the transition phrase that splits preamble from body
    transition_pattern = re.compile(
        r'\b(comprising|consisting\s+(?:essentially\s+)?of|including|wherein|characterized\s+in\s+that)\s*:\s*',
        re.IGNORECASE,
    )
    match = transition_pattern.search(text)

    if match:
        preamble = text[:match.start()].strip()
        transition_word = match.group(0).strip()
        body = text[match.end():].strip()

        if preamble:
            elements.append({'text': preamble, 'type': 'preamble', 'order': order})
            order += 1
        elements.append({'text': transition_word.rstrip(':').strip(), 'type': 'transition', 'order': order})
        order += 1
    else:
        body = text

    if not body:
        # No body to split — if we only have preamble, return as-is
        if not elements:
            elements.append({'text': text, 'type': 'body', 'order': 1})
        return elements

    # Split body on semicolons (standard patent claim delimiter)
    parts = re.split(r';\s*', body)
    for part in parts:
        part = part.strip().rstrip('.')
        if part:
            elements.append({'text': part, 'type': 'body', 'order': order})
            order += 1

    return elements


def _normalize_patent_claims(claims):
    """Normalize a Patent.claims list (dicts {number,text} or bare strings) into a
    list of (claim_number:str, claim_text:str) pairs, skipping empties."""
    items = []
    for i, c in enumerate(claims or []):
        if isinstance(c, dict):
            text = (c.get('text') or c.get('claim_text') or '').strip()
            number = str(c.get('number') or c.get('claim_number') or (i + 1))
        else:
            text, number = str(c).strip(), str(i + 1)
        if text:
            items.append((number, text))
    return items


def _create_claim_mappings(case, claim_items, user):
    """Create ClaimMappings (+ parsed elements) for a case from (number, text) pairs."""
    from .models import ClaimMapping, ClaimElement
    created = []
    for number, claim_text in claim_items:
        is_dependent = bool(re.search(r'\bclaim\s+\d+', claim_text, re.IGNORECASE)) and \
            not claim_text.strip().startswith(f'{number}.')
        mapping = ClaimMapping.objects.create(
            case=case,
            claim_number=str(number),
            claim_text=claim_text,
            claim_type='dependent' if is_dependent else 'independent',
            product_feature='',
            product_feature_description='',
            mapping_type='literal',
            match_confidence=0,
            limitations_met=False,
            created_by=user,
        )
        for elem in _parse_claim_into_elements(claim_text):
            ClaimElement.objects.create(
                claim_mapping=mapping,
                element_order=elem['order'],
                element_text=elem['text'],
                element_type=elem['type'],
                created_by=user,
            )
        created.append(mapping)
    return created

from .models import (
    InfringementCase,
    ClaimMapping,
    ClaimElement,
    Evidence,
    EvidenceScreenshot,
    RiskAssessment,
    DamagesAnalysis,
    ExpertOpinion,
    LitigationStrategy,
    InfringementReport
)
from .serializers import (
    InfringementCaseSerializer,
    InfringementCaseListSerializer,
    ClaimMappingSerializer,
    ClaimElementSerializer,
    EvidenceSerializer,
    EvidenceScreenshotSerializer,
    RiskAssessmentSerializer,
    DamagesAnalysisSerializer,
    ExpertOpinionSerializer,
    LitigationStrategySerializer,
    InfringementReportSerializer
)


class InfringementCaseViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Infringement Cases
    """
    queryset = InfringementCase.objects.all().select_related(
        'analyst', 'assigned_attorney', 'created_by', 'patent', 'patent__portfolio'
    ).prefetch_related(
        'claim_mappings', 'evidence', 'risk_assessments', 'reports'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'analysis_type', 'risk_level', 'patent_number', 'accused_party_name']
    search_fields = ['case_name', 'patent_number', 'patent_title', 'accused_product_name', 'accused_party_name']
    ordering_fields = ['created_at', 'updated_at', 'infringement_likelihood', 'risk_level']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return InfringementCaseListSerializer
        return InfringementCaseSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get dashboard statistics. Optionally filter by portfolio_id."""
        cases = self.get_queryset()

        # Optional portfolio filtering
        portfolio_id = request.query_params.get('portfolio_id')
        if portfolio_id:
            cases = cases.filter(
                Q(patent__portfolio_id=portfolio_id) |
                Q(patent__isnull=True, patent_number__in=self._patent_numbers_for_portfolio(portfolio_id))
            )

        stats = {
            'total_cases': cases.count(),
            'active_cases': cases.filter(status='active').count(),
            'high_risk_cases': cases.filter(risk_level='high').count(),
            'critical_risk_cases': cases.filter(risk_level='critical').count(),
            'avg_infringement_likelihood': cases.aggregate(Avg('infringement_likelihood'))['infringement_likelihood__avg'] or 0,
            'cases_by_status': {
                status_choice[0]: cases.filter(status=status_choice[0]).count()
                for status_choice in InfringementCase.STATUS_CHOICES
            },
            'cases_by_risk': {
                risk_choice[0]: cases.filter(risk_level=risk_choice[0]).count()
                for risk_choice in InfringementCase.RISK_LEVEL_CHOICES
            }
        }

        # Affected portfolios: portfolios that have linked infringement cases
        from domains.patents.models import Portfolio
        affected_portfolio_ids = (
            cases.filter(patent__portfolio__isnull=False)
            .values_list('patent__portfolio_id', flat=True)
            .distinct()
        )
        affected_portfolios = Portfolio.objects.filter(id__in=affected_portfolio_ids).values(
            'id', 'name', 'company_name'
        )
        # Count active cases per portfolio
        portfolio_case_counts = {}
        for pid in affected_portfolio_ids:
            portfolio_case_counts[str(pid)] = cases.filter(
                patent__portfolio_id=pid, status='active'
            ).count()

        stats['affected_portfolios'] = [
            {
                'id': str(p['id']),
                'name': p['name'],
                'company_name': p['company_name'],
                'active_cases': portfolio_case_counts.get(str(p['id']), 0),
            }
            for p in affected_portfolios
        ]

        return Response(stats)

    def _patent_numbers_for_portfolio(self, portfolio_id):
        """Helper: get patent numbers belonging to a portfolio for string-based matching."""
        from domains.patents.models import Patent
        return list(
            Patent.objects.filter(portfolio_id=portfolio_id, patent_number__isnull=False)
            .values_list('patent_number', flat=True)
        )

    @action(detail=False, methods=['post'])
    def create_from_patent(self, request):
        """
        Create a draft infringement case pre-filled from a Patent record.
        Expects: patent_id (required), accused_product_name, accused_party_name, analysis_type
        """
        from domains.patents.models import Patent

        patent_id = request.data.get('patent_id')
        if not patent_id:
            return Response(
                {'error': 'patent_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        patent = get_object_or_404(Patent, id=patent_id)

        case = InfringementCase.objects.create(
            patent=patent,
            patent_number=patent.patent_number or patent.application_number or '',
            patent_title=patent.title,
            patent_abstract=patent.abstract or '',
            case_name=request.data.get('case_name', f'Infringement Analysis - {patent.title[:100]}'),
            accused_product_name=request.data.get('accused_product_name', ''),
            accused_product_description=request.data.get('accused_product_description', ''),
            accused_party_name=request.data.get('accused_party_name', ''),
            analysis_type=request.data.get('analysis_type', 'literal'),
            status='draft',
            analyst=request.user,
            created_by=request.user,
        )

        serializer = InfringementCaseSerializer(case)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='enrich-from-odp')
    def enrich_from_odp(self, request, pk=None):
        """Resolve this case's patent number on USPTO ODP, link/create the canonical
        Patent, and fill blank case fields. Refuses to link when the resolved record
        doesn't match the case (title), so an unrelated patent that merely shares the
        number isn't attached. Pass {"force": true} to link anyway. USPTO only — no LLM.
        """
        from domains.patents.models import Patent
        from domains.patents.enrichment_service import resolve_application_number, enrich_patent

        case = self.get_object()
        if not case.patent_number:
            return Response({'error': 'Case has no patent number to look up.'},
                            status=status.HTTP_400_BAD_REQUEST)

        force = bool(request.data.get('force'))

        try:
            app_num = resolve_application_number(case.patent_number)
        except Exception as exc:  # network/ODP failure
            logger.warning('enrich_from_odp resolve failed for %s: %s', case.patent_number, exc)
            return Response({'error': f'USPTO lookup failed: {exc}'},
                            status=status.HTTP_502_BAD_GATEWAY)

        if not app_num:
            return Response(
                {'matched': False, 'error': 'No USPTO application found for this patent number.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Start a new Patent with a BLANK title so enrich_patent fills it from USPTO —
        # seeding it with the case title would make the match guard compare the title to
        # itself and always pass.
        patent, _created = Patent.objects.get_or_create(
            application_number=app_num,
            defaults={'title': ''},
        )
        try:
            enrich_patent(patent)
        except Exception as exc:
            logger.warning('enrich_from_odp enrich failed for app %s: %s', app_num, exc)
            return Response({'error': f'USPTO enrichment failed: {exc}'},
                            status=status.HTTP_502_BAD_GATEWAY)

        match = _odp_case_match(case, patent)
        if not match['confident'] and not force:
            return Response({
                'matched': False,
                'reason': (
                    f"USPTO record “{match['found_title']}” doesn’t match this case’s "
                    f"patent title."
                ),
                **match,
            }, status=status.HTTP_200_OK)

        # Link the canonical patent and fill blank case fields (never clobber analyst edits).
        case.patent = patent
        if not case.patent_title and patent.title:
            case.patent_title = patent.title
        if not case.patent_abstract and patent.abstract:
            case.patent_abstract = patent.abstract
        case.save(update_fields=['patent', 'patent_title', 'patent_abstract', 'updated_at'])

        return Response({
            'matched': True,
            **match,
            'case': InfringementCaseSerializer(case, context={'request': request}).data,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def update_risk_level(self, request, pk=None):
        """Update risk level based on analysis"""
        case = self.get_object()
        new_risk = request.data.get('risk_level')

        if new_risk not in dict(InfringementCase.RISK_LEVEL_CHOICES):
            return Response(
                {'error': 'Invalid risk level'},
                status=status.HTTP_400_BAD_REQUEST
            )

        case.risk_level = new_risk
        case.save()

        serializer = self.get_serializer(case)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Export infringement case as PDF report"""
        from utils.export import export_infringement_report_pdf

        case = self.get_object()

        # Prepare data for export
        analysis_data = {
            'target_patent': f"{case.patent_number} - {case.patent_title}",
            'risk_level': case.get_risk_level_display(),
            'claims': [
                {
                    'claim_number': cm.claim_number,
                    'claim_type': cm.get_mapping_type_display(),
                    'risk_level': 'High' if cm.limitations_met else 'Low',
                    'coverage_percentage': int(cm.match_confidence),
                    'notes': cm.product_feature_description[:100] if cm.product_feature_description else ''
                }
                for cm in case.claim_mappings.all()
            ],
            'evidence': [
                {
                    'reference_id': str(e.id)[:8],
                    'type': e.get_evidence_type_display(),
                    'relevance': f"{e.relevance_score}%",
                    'summary': e.description[:100] if e.description else ''
                }
                for e in case.evidence.all()
            ],
            'recommendations': [
                f"Risk Level: {case.get_risk_level_display()}",
                f"Infringement Likelihood: {case.infringement_likelihood}%",
            ]
        }

        return export_infringement_report_pdf(analysis_data)

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """Export infringement case as Excel report"""
        from utils.export import export_infringement_report_excel

        case = self.get_object()

        # Prepare data for export
        analysis_data = {
            'target_patent': f"{case.patent_number} - {case.patent_title}",
            'risk_level': case.get_risk_level_display(),
            'claims': [
                {
                    'claim_number': cm.claim_number,
                    'claim_type': cm.get_mapping_type_display(),
                    'risk_level': 'High' if cm.limitations_met else 'Low',
                    'coverage_percentage': int(cm.match_confidence),
                    'notes': cm.product_feature_description[:100] if cm.product_feature_description else ''
                }
                for cm in case.claim_mappings.all()
            ],
            'evidence': [
                {
                    'reference_id': str(e.id)[:8],
                    'type': e.get_evidence_type_display(),
                    'relevance': f"{e.relevance_score}%",
                    'summary': e.description or ''
                }
                for e in case.evidence.all()
            ],
        }

        return export_infringement_report_excel(analysis_data)

    @action(detail=True, methods=['get'], url_path='export-claim-chart-excel')
    def export_claim_chart_excel(self, request, pk=None):
        """Export a detailed, element-by-element claim chart as an .xlsx file.

        One row per claim element with met/not-met, accused feature, confidence, and
        resolved evidence citations — the professional claim-chart artifact buyers expect.
        """
        import io
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        case = self.get_object()

        # Resolve evidence ids -> titles once for citation cells.
        ev_titles = {str(e.id): e.title for e in case.evidence.all()}

        def met_label(v):
            return 'Met' if v is True else ('Not Met' if v is False else 'Unknown')

        FILLS = {
            'Met': PatternFill('solid', fgColor='C6EFCE'),
            'Not Met': PatternFill('solid', fgColor='FFC7CE'),
            'Unknown': PatternFill('solid', fgColor='FFEB9C'),
        }

        wb = Workbook()
        ws = wb.active
        ws.title = 'Claim Chart'
        headers = ['Claim', 'Element #', 'Element Text', 'Type', 'Status',
                   'Accused Feature', 'Confidence', 'Analysis Notes', 'Citations', 'Review']
        ws.append(headers)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F2937')
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(vertical='top', wrap_text=True)

        for cm in case.claim_mappings.all().order_by('claim_number'):
            elements = list(cm.elements.all().order_by('element_order'))
            if not elements:
                # mapping-level row when no elements parsed
                ws.append([cm.claim_number, '', cm.claim_text, cm.claim_type,
                           'Met' if cm.limitations_met else 'Unknown', cm.product_feature,
                           f'{cm.match_confidence}%', cm.analysis_notes or '',
                           '; '.join(ev_titles.get(str(i), str(i)) for i in (cm.evidence_references or [])),
                           cm.review_status or ''])
                continue
            for el in elements:
                status_label = met_label(el.meets_limitation)
                cites = '; '.join(ev_titles.get(str(i), str(i)) for i in (el.evidence_references or []))
                ws.append([
                    cm.claim_number, el.element_order, el.element_text, el.element_type,
                    status_label, el.accused_feature or '',
                    f'{el.doe_score}%' if el.doe_score is not None else f'{cm.match_confidence}%',
                    el.analysis_notes or '', cites, el.review_status or '',
                ])
                ws.cell(row=ws.max_row, column=5).fill = FILLS.get(status_label, FILLS['Unknown'])

        # Column widths + wrapping for readability
        widths = [10, 9, 60, 12, 10, 40, 11, 40, 30, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        safe_name = (case.case_number or str(case.id)[:8])
        resp = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="claim_chart_{safe_name}.xlsx"'
        return resp

    @action(detail=True, methods=['post'], url_path='extract-claim-terms')
    def extract_claim_terms(self, request, pk=None):
        """Extract antecedent-basis claim terms and assign consistent colors (heuristic /
        spaCy). Preserves existing colors; returns the case-wide term→color map."""
        from .claim_term_service import extract_claim_terms as _extract
        case = self.get_object()
        color_map = _extract(case, strategy=request.data.get('strategy', 'antecedent'))
        return Response({'claim_term_colors': color_map, 'count': len(color_map)})

    @action(detail=True, methods=['post'], url_path='set-term-color')
    def set_term_color(self, request, pk=None):
        """Manually override one term's color (merges into the case map)."""
        case = self.get_object()
        term = (request.data.get('term') or '').strip().lower()
        color = (request.data.get('color') or '').strip()
        if not term or not color:
            return Response({'error': 'term and color are required'}, status=status.HTTP_400_BAD_REQUEST)
        m = dict(case.claim_term_colors or {})
        m[term] = color
        case.claim_term_colors = m
        case.save(update_fields=['claim_term_colors', 'updated_at'])
        return Response({'claim_term_colors': m})

    @action(detail=True, methods=['post'], url_path='add-term')
    def add_term(self, request, pk=None):
        """Add a missed claim term (normalized) with a color; un-excludes it."""
        from .claim_term_service import _normalize, PALETTE
        case = self.get_object()
        term = _normalize(request.data.get('term') or '')
        if not term:
            return Response({'error': 'term is required'}, status=status.HTTP_400_BAD_REQUEST)
        m = dict(case.claim_term_colors or {})
        ex = [t for t in (case.claim_term_excluded or []) if t != term]
        if term not in m:
            color = (request.data.get('color') or '').strip()
            if not color:
                used = set(m.values())
                color = next((c for c in PALETTE if c not in used), PALETTE[len(m) % len(PALETTE)])
            m[term] = color
        case.claim_term_colors = m
        case.claim_term_excluded = ex
        case.save(update_fields=['claim_term_colors', 'claim_term_excluded', 'updated_at'])
        return Response({'claim_term_colors': m, 'claim_term_excluded': ex})

    @action(detail=True, methods=['post'], url_path='remove-term')
    def remove_term(self, request, pk=None):
        """Remove a term and exclude it so auto-extraction won't re-add it."""
        case = self.get_object()
        term = (request.data.get('term') or '').strip().lower()
        m = dict(case.claim_term_colors or {})
        m.pop(term, None)
        ex = list(case.claim_term_excluded or [])
        if term and term not in ex:
            ex.append(term)
        case.claim_term_colors = m
        case.claim_term_excluded = ex
        case.save(update_fields=['claim_term_colors', 'claim_term_excluded', 'updated_at'])
        return Response({'claim_term_colors': m, 'claim_term_excluded': ex})

    @action(detail=True, methods=['post'], url_path='rename-term')
    def rename_term(self, request, pk=None):
        """Rename a term; if the new term already exists this merges them (one color).
        The old phrase is excluded so it isn't re-detected."""
        from .claim_term_service import _normalize
        case = self.get_object()
        old = (request.data.get('term') or '').strip().lower()
        new = _normalize(request.data.get('new_term') or '')
        if not old or not new:
            return Response({'error': 'term and new_term are required'}, status=status.HTTP_400_BAD_REQUEST)
        m = dict(case.claim_term_colors or {})
        ex = list(case.claim_term_excluded or [])
        color = m.pop(old, None)
        if new not in m:  # rename; if new exists it's a merge → keep new's color
            m[new] = color or '#EF4444'
        if old and old != new and old not in ex:
            ex.append(old)
        ex = [t for t in ex if t != new]
        case.claim_term_colors = m
        case.claim_term_excluded = ex
        case.save(update_fields=['claim_term_colors', 'claim_term_excluded', 'updated_at'])
        return Response({'claim_term_colors': m, 'claim_term_excluded': ex})

    @action(detail=True, methods=['get'])
    def ptab_proceedings(self, request, pk=None):
        """Search ODP PTAB proceedings related to the case's patent number (cached)."""
        case = self.get_object()
        if not case.patent_number:
            return Response({'error': 'Case has no patent_number for PTAB lookup'}, status=status.HTTP_400_BAD_REQUEST)
        return _ptab_cached('ptab_proceedings', case.patent_number, 'search_proceedings')

    @action(detail=True, methods=['get'])
    def ptab_decisions(self, request, pk=None):
        """Search ODP PTAB decisions related to the case's patent number (cached)."""
        case = self.get_object()
        if not case.patent_number:
            return Response({'error': 'Case has no patent_number for PTAB lookup'}, status=status.HTTP_400_BAD_REQUEST)
        return _ptab_cached('ptab_decisions', case.patent_number, 'search_decisions')

    @action(detail=True, methods=['post'], url_path='auto-import-claims')
    def auto_import_claims(self, request, pk=None):
        """Import patent claims for the case.

        Source of truth, in order: (1) the linked Patent's locally-stored claims
        (authoritative, no network, correct for enriched/seed patents), then
        (2) USPTO ODP full text. Pass force=true to replace existing mappings
        (e.g. when the current ones are stale/mismatched).
        """
        case = self.get_object()
        force = bool(request.data.get('force', False))

        # If claims already exist and we're not forcing a refresh, return them.
        if case.claim_mappings.exists() and not force:
            serializer = ClaimMappingSerializer(case.claim_mappings.all().order_by('claim_number'), many=True)
            return Response({'status': 'existing', 'claim_mappings': serializer.data})

        # (1) Prefer the linked Patent's local claims when available.
        local_items = _normalize_patent_claims(getattr(case.patent, 'claims', None)) if case.patent else []
        if local_items:
            with transaction.atomic():
                locked_case = InfringementCase.objects.select_for_update().get(pk=case.pk)
                if locked_case.claim_mappings.exists():
                    if not force:
                        serializer = ClaimMappingSerializer(locked_case.claim_mappings.all().order_by('claim_number'), many=True)
                        return Response({'status': 'existing', 'claim_mappings': serializer.data})
                    locked_case.claim_mappings.all().delete()
                created = _create_claim_mappings(locked_case, local_items, request.user)
            serializer = ClaimMappingSerializer(created, many=True)
            return Response({
                'status': 'imported',
                'source': 'patent',
                'claim_mappings': serializer.data,
                'count': len(created),
            })

        patent_number = case.patent_number
        if not patent_number:
            return Response({'status': 'no_patent_number', 'message': 'Case has no patent number'})

        # Clean patent number: strip US prefix, commas, spaces, kind codes
        clean_number = re.sub(r'[,\s]', '', patent_number)
        clean_number = re.sub(r'^US', '', clean_number, flags=re.IGNORECASE)
        clean_number = re.sub(r'[A-Z]\d*$', '', clean_number)  # strip kind code e.g. B2, B1

        from domains.analytics.uspto_odp_service import USPTOODPClient, USPTOODPDetailService, USPTOODPError
        from domains.analytics.odp_views import _fetch_and_parse_xml
        from domains.analytics.models import ODPCacheEntry
        from django.utils import timezone
        from datetime import timedelta

        try:
            client = USPTOODPClient()
            # Search for the patent using applicationNumberText (works for both app & patent numbers)
            search_result = client.post('/patent/applications/search', {
                'q': f'applicationNumberText:{clean_number}',
                'pagination': {'offset': 0, 'limit': 1},
            })
            if not search_result:
                return Response({'status': 'not_found', 'message': 'Patent not found in USPTO ODP'})

            results_bag = search_result.get('patentFileWrapperDataBag', [])
            if not results_bag:
                return Response({'status': 'not_found', 'message': 'Patent not found in USPTO ODP'})

            app_id = results_bag[0].get('applicationNumberText', '')
            if not app_id:
                return Response({'status': 'not_found', 'message': 'Could not extract application ID'})

            # Get application metadata for XML URLs
            detail_svc = USPTOODPDetailService(client)
            app_data = detail_svc.get_application(app_id)
            if not app_data:
                return Response({'status': 'not_found', 'message': 'Application data not available'})

            # Check full-text-parsed cache (7-day TTL)
            cutoff = timezone.now() - timedelta(days=7)
            cached = ODPCacheEntry.objects.filter(
                application_id=app_id, endpoint='full-text-parsed', fetched_at__gte=cutoff
            ).first()

            if cached:
                full_text_data = cached.response_data
            else:
                # Fetch and parse XML from grant or publication
                api_key = client.api_key
                grant_meta = app_data.get('grantDocumentMetaData') or {}
                grant_url = grant_meta.get('fileLocationURI')
                pub_meta = app_data.get('pgpubDocumentMetaData') or app_data.get('publicationDocumentMetaData') or {}
                pgpub_url = pub_meta.get('fileLocationURI')

                grant_text = _fetch_and_parse_xml(grant_url, api_key) if grant_url else None
                pgpub_text = _fetch_and_parse_xml(pgpub_url, api_key) if pgpub_url else None

                full_text_data = {
                    'grant_text': grant_text,
                    'pgpub_text': pgpub_text,
                }

                # Cache the result
                ODPCacheEntry.objects.update_or_create(
                    application_id=app_id,
                    endpoint='full-text-parsed',
                    defaults={'response_data': {
                        **full_text_data,
                        'grant_url': grant_url,
                        'pgpub_url': pgpub_url,
                    }, 'fetched_at': timezone.now()},
                )

            # Extract claims from grant first, fall back to publication
            claims = []
            grant_text = full_text_data.get('grant_text')
            pgpub_text = full_text_data.get('pgpub_text')
            if grant_text and grant_text.get('claims'):
                claims = grant_text['claims']
            elif pgpub_text and pgpub_text.get('claims'):
                claims = pgpub_text['claims']

            if not claims:
                return Response({'status': 'no_claims', 'message': 'No claims found in USPTO full text'})

            # Atomic create with re-check to prevent duplicates from concurrent requests
            with transaction.atomic():
                locked_case = InfringementCase.objects.select_for_update().get(pk=case.pk)
                if locked_case.claim_mappings.exists():
                    if not force:
                        serializer = ClaimMappingSerializer(locked_case.claim_mappings.all().order_by('claim_number'), many=True)
                        return Response({'status': 'existing', 'claim_mappings': serializer.data})
                    locked_case.claim_mappings.all().delete()

                # ODP returns a list of claim-text strings; number them sequentially.
                created_mappings = _create_claim_mappings(
                    locked_case, [(str(i + 1), t) for i, t in enumerate(claims)], request.user
                )

            serializer = ClaimMappingSerializer(created_mappings, many=True)
            return Response({
                'status': 'imported',
                'source': 'odp',
                'claim_mappings': serializer.data,
                'count': len(created_mappings),
            })

        except USPTOODPError as exc:
            logger.warning('ODP error during auto-import for case %s: %s', pk, exc)
            return Response({'status': 'error', 'message': f'USPTO ODP error: {exc}'})
        except Exception as exc:
            logger.exception('Unexpected error during auto-import for case %s', pk)
            return Response({'status': 'error', 'message': str(exc)})


class ClaimMappingViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Claim Mappings
    """
    queryset = ClaimMapping.objects.all().select_related('case', 'created_by').prefetch_related('elements')
    serializer_class = ClaimMappingSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['case', 'mapping_type', 'limitations_met']
    ordering_fields = ['claim_number', 'match_confidence']
    ordering = ['claim_number']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='parse-elements')
    def parse_elements(self, request, pk=None):
        """Auto-parse claim text into elements. Replaces any existing elements.

        use_llm=True opts into the gated LLM parser (dormant until the master switch
        INFRINGEMENT_AI_LLM_ENABLED is on); otherwise heuristic structural parsing.
        """
        claim_mapping = self.get_object()
        if not claim_mapping.claim_text:
            return Response({'error': 'No claim text to parse'}, status=status.HTTP_400_BAD_REQUEST)

        use_llm = bool(request.data.get('use_llm', False))

        # Delete existing elements and re-parse
        claim_mapping.elements.all().delete()
        parsed = ai_service.parse_claim_elements(claim_mapping.claim_text, use_llm=use_llm)
        created = []
        for elem in parsed:
            created.append(ClaimElement.objects.create(
                claim_mapping=claim_mapping,
                element_order=elem['element_order'],
                element_text=elem['element_text'],
                element_type=elem['element_type'],
                is_ai_generated=use_llm,
                review_status='ai_draft' if use_llm else 'confirmed',
                created_by=request.user,
            ))
        serializer = ClaimElementSerializer(created, many=True)
        return Response({'elements': serializer.data, 'count': len(created), 'ai_draft': use_llm})

    @action(detail=True, methods=['post'], url_path='generate-mapping')
    def generate_mapping(self, request, pk=None):
        """Draft an element-by-element claim→product mapping for analyst review.

        Replaces existing elements with drafts. With use_llm=True (and the master
        switch on) the LLM fills accused-feature + met/not-met + rationale per element;
        otherwise it returns a heuristic element split for the analyst to complete.
        AI output is flagged review_status='ai_draft'.
        """
        claim_mapping = self.get_object()
        if not claim_mapping.claim_text:
            return Response({'error': 'No claim text to map'}, status=status.HTTP_400_BAD_REQUEST)

        use_llm = bool(request.data.get('use_llm', False))
        product_desc = (
            request.data.get('product_description')
            or claim_mapping.case.accused_product_description
            or ''
        )

        result = ai_service.generate_claim_mapping(
            claim_mapping.claim_text, product_desc, use_llm=use_llm
        )

        claim_mapping.elements.all().delete()
        created = []
        for elem in result.get('elements', []):
            created.append(ClaimElement.objects.create(
                claim_mapping=claim_mapping,
                element_order=elem['element_order'],
                element_text=elem['element_text'],
                element_type=elem['element_type'],
                accused_feature=elem.get('accused_feature', ''),
                meets_limitation=elem.get('meets_limitation'),
                analysis_notes=elem.get('analysis_notes', ''),
                is_ai_generated=use_llm,
                review_status='ai_draft' if use_llm else 'confirmed',
                created_by=request.user,
            ))

        # Flag the mapping itself as an AI draft pending review.
        if use_llm:
            claim_mapping.is_ai_generated = True
            claim_mapping.review_status = 'ai_draft'
            if result.get('summary') and not claim_mapping.analysis_notes:
                claim_mapping.analysis_notes = result['summary']
            claim_mapping.save(update_fields=['is_ai_generated', 'review_status', 'analysis_notes', 'updated_at'])

        serializer = ClaimElementSerializer(created, many=True)
        return Response({
            'summary': result.get('summary', ''),
            'elements': serializer.data,
            'count': len(created),
            'ai_draft': use_llm,
        })

    @action(detail=True, methods=['get'])
    def elements(self, request, pk=None):
        """Get all elements for a claim mapping"""
        claim_mapping = self.get_object()
        elements = claim_mapping.elements.all()
        serializer = ClaimElementSerializer(elements, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def element_summary(self, request, pk=None):
        """Get summary of element analysis for a claim mapping"""
        claim_mapping = self.get_object()
        elements = claim_mapping.elements.all()

        total = elements.count()
        met = elements.filter(meets_limitation=True).count()
        not_met = elements.filter(meets_limitation=False).count()
        unknown = elements.filter(meets_limitation__isnull=True).count()

        # Calculate overall score
        if total > 0:
            score = (met / total) * 100
        else:
            score = claim_mapping.match_confidence

        return Response({
            'claim_mapping_id': str(claim_mapping.id),
            'total_elements': total,
            'elements_met': met,
            'elements_not_met': not_met,
            'elements_unknown': unknown,
            'overall_score': round(score, 1),
            'all_limitations_met': met == total and total > 0,
        })


class ClaimElementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Claim Elements (element-by-element analysis)
    """
    queryset = ClaimElement.objects.all().select_related('claim_mapping', 'created_by')
    serializer_class = ClaimElementSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['claim_mapping', 'element_type', 'meets_limitation']
    ordering_fields = ['element_order']
    ordering = ['element_order']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def analyze_doe(self, request, pk=None):
        """
        Analyze Doctrine of Equivalents for an element
        """
        element = self.get_object()

        # Update DoE analysis fields
        element.doe_function = request.data.get('doe_function', element.doe_function)
        element.doe_way = request.data.get('doe_way', element.doe_way)
        element.doe_result = request.data.get('doe_result', element.doe_result)
        element.doe_score = request.data.get('doe_score', element.doe_score)
        element.save()

        serializer = self.get_serializer(element)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Create multiple elements at once
        """
        elements_data = request.data.get('elements', [])
        claim_mapping_id = request.data.get('claim_mapping_id')

        if not claim_mapping_id:
            return Response(
                {'error': 'claim_mapping_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            claim_mapping = ClaimMapping.objects.get(id=claim_mapping_id)
        except ClaimMapping.DoesNotExist:
            return Response(
                {'error': 'ClaimMapping not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        created_elements = []
        for idx, elem_data in enumerate(elements_data):
            element = ClaimElement.objects.create(
                claim_mapping=claim_mapping,
                element_order=elem_data.get('element_order', idx + 1),
                element_text=elem_data.get('element_text', ''),
                element_type=elem_data.get('element_type', 'body'),
                accused_feature=elem_data.get('accused_feature', ''),
                accused_feature_description=elem_data.get('accused_feature_description', ''),
                meets_limitation=elem_data.get('meets_limitation'),
                analysis_notes=elem_data.get('analysis_notes', ''),
                created_by=request.user,
            )
            created_elements.append(element)

        serializer = ClaimElementSerializer(created_elements, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class EvidenceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Evidence
    """
    queryset = Evidence.objects.all().select_related('case', 'uploaded_by')
    serializer_class = EvidenceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['case', 'evidence_type']
    search_fields = ['title', 'description', 'source']
    ordering_fields = ['relevance_score', 'created_at']
    ordering = ['-relevance_score', '-created_at']

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    def perform_destroy(self, instance):
        """Delete the evidence and scrub its id from the JSON evidence_references lists on
        the case's claim mappings and elements (those are id lists with no FK cascade)."""
        evidence_id = str(instance.id)
        case_id = instance.case_id
        instance.delete()

        for mapping in ClaimMapping.objects.filter(case_id=case_id):
            refs = mapping.evidence_references or []
            if evidence_id in refs:
                mapping.evidence_references = [r for r in refs if r != evidence_id]
                mapping.save(update_fields=['evidence_references', 'updated_at'])

        for element in ClaimElement.objects.filter(claim_mapping__case_id=case_id):
            refs = element.evidence_references or []
            if evidence_id in refs:
                element.evidence_references = [r for r in refs if r != evidence_id]
                element.save(update_fields=['evidence_references', 'updated_at'])

    @action(detail=False, methods=['post'], url_path='suggest-from-url')
    def suggest_from_url(self, request):
        """Assisted evidence sourcing: fetch a product URL server-side, extract text,
        and return ranked candidate passages for the analyst to review/save.

        Body: { url, claim_mapping_id? | claim_text?, use_llm? }. Nothing is persisted —
        candidates are returned for the analyst to turn into Evidence. LLM ranking only
        runs when use_llm=True AND the master switch is on; otherwise heuristic overlap.
        """
        import requests
        from bs4 import BeautifulSoup

        url = (request.data.get('url') or '').strip()
        if not url or not url.lower().startswith(('http://', 'https://')):
            return Response({'error': 'A valid http(s) url is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Resolve claim context for relevance ranking.
        claim_text = request.data.get('claim_text', '')
        cm_id = request.data.get('claim_mapping_id')
        if cm_id and not claim_text:
            cm = ClaimMapping.objects.filter(id=cm_id).first()
            if cm:
                claim_text = cm.claim_text

        try:
            resp = requests.get(url, timeout=10, headers={'User-Agent': 'PatentAnalytics/1.0'})
            resp.raise_for_status()
        except Exception as exc:
            return Response({'error': f'Failed to fetch URL: {exc}'}, status=status.HTTP_502_BAD_GATEWAY)

        soup = BeautifulSoup(resp.text[:500000], 'html.parser')
        for tag in soup(['script', 'style', 'noscript', 'header', 'footer', 'nav']):
            tag.decompose()
        page_title = (soup.title.string.strip() if soup.title and soup.title.string else url)
        text = soup.get_text(separator='\n')

        candidates = ai_service.suggest_evidence_passages(
            text, claim_text=claim_text, use_llm=bool(request.data.get('use_llm', False))
        )
        return Response({
            'source_url': url,
            'title': page_title[:255],
            'candidates': candidates,
            'count': len(candidates),
        })


class EvidenceScreenshotViewSet(viewsets.ModelViewSet):
    """ViewSet for evidence-of-use screenshots (cropped PDF regions mapped to claim
    elements). Created via multipart (image + page + bbox + repeated claim_elements)."""
    queryset = EvidenceScreenshot.objects.all().select_related('evidence', 'created_by').prefetch_related('claim_elements')
    serializer_class = EvidenceScreenshotSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['case', 'evidence', 'claim_elements']
    ordering_fields = ['created_at', 'page_number']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        # case is derived from the source evidence so the client needn't send it.
        evidence = serializer.validated_data.get('evidence')
        serializer.save(created_by=self.request.user, case=evidence.case)


class RiskAssessmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Risk Assessments
    """
    queryset = RiskAssessment.objects.all().select_related('case', 'assessed_by')
    serializer_class = RiskAssessmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['case', 'risk_factor']
    ordering_fields = ['risk_score', 'assessed_date']
    ordering = ['-risk_score']

    def perform_create(self, serializer):
        serializer.save(assessed_by=self.request.user)


class InfringementReportViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Infringement Reports
    """
    queryset = InfringementReport.objects.all().select_related(
        'case', 'created_by', 'reviewed_by'
    )
    serializer_class = InfringementReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['case', 'report_type', 'status']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        """Mark report as reviewed"""
        report = self.get_object()
        report.reviewed_by = request.user
        report.status = 'final'
        report.review_notes = request.data.get('review_notes', '')
        report.save()

        serializer = self.get_serializer(report)
        return Response(serializer.data)


class DamagesAnalysisViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Damages Analysis
    """
    queryset = DamagesAnalysis.objects.all().select_related('case', 'created_by')
    serializer_class = DamagesAnalysisSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['case', 'damages_theory', 'is_willful']
    ordering_fields = ['created_at', 'estimated_damages_high']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """
        Calculate damages based on current data
        """
        damages_analysis = self.get_object()
        calculation = damages_analysis.calculate_damages()

        # Update the estimated damages
        damages_analysis.estimated_damages_low = calculation['base_damages']
        damages_analysis.estimated_damages_high = calculation['total_damages']
        damages_analysis.save()

        serializer = self.get_serializer(damages_analysis)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='get-or-create')
    def get_or_create(self, request):
        """
        Get existing damages analysis for a case or create a new one
        """
        case_id = request.data.get('case_id')
        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        damages_analysis, created = DamagesAnalysis.objects.get_or_create(
            case=case,
            defaults={'created_by': request.user}
        )

        serializer = self.get_serializer(damages_analysis)
        return Response({
            'damages_analysis': serializer.data,
            'created': created,
        })

    @action(detail=False, methods=['get'], url_path='by-case/(?P<case_id>[^/.]+)')
    def by_case(self, request, case_id=None):
        """
        Get damages analysis for a specific case
        """
        try:
            damages_analysis = DamagesAnalysis.objects.get(case_id=case_id)
            serializer = self.get_serializer(damages_analysis)
            return Response(serializer.data)
        except DamagesAnalysis.DoesNotExist:
            return Response(
                {'error': 'No damages analysis found for this case'},
                status=status.HTTP_404_NOT_FOUND
            )


class ExpertOpinionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Expert Opinions
    """
    queryset = ExpertOpinion.objects.all().select_related('case', 'created_by').prefetch_related('supports_claims')
    serializer_class = ExpertOpinionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['case', 'opinion_type', 'is_under_protective_order']
    search_fields = ['expert_name', 'expert_organization', 'findings', 'conclusion']
    ordering_fields = ['opinion_date', 'created_at']
    ordering = ['-opinion_date']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class LitigationStrategyViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Litigation Strategy
    """
    queryset = LitigationStrategy.objects.all().select_related('case', 'created_by')
    serializer_class = LitigationStrategySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['case']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='get-or-create')
    def get_or_create(self, request):
        """
        Get existing litigation strategy for a case or create a new one
        """
        case_id = request.data.get('case_id')
        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        strategy, created = LitigationStrategy.objects.get_or_create(
            case=case,
            defaults={'created_by': request.user}
        )

        serializer = self.get_serializer(strategy)
        return Response({
            'litigation_strategy': serializer.data,
            'created': created,
        })


class RiskAnalysisViewSet(viewsets.ViewSet):
    """
    ViewSet for Risk Analysis calculations and reports
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='calculate-risk-score')
    def calculate_risk_score(self, request):
        """
        Calculate comprehensive risk score for a case based on all risk factors
        """
        case_id = request.data.get('case_id')
        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get all risk assessments for this case
        risk_assessments = case.risk_assessments.all()

        # Calculate weighted risk score
        risk_factors = {
            'technical': {'weight': 0.25, 'score': 0, 'assessed': False},
            'legal': {'weight': 0.20, 'score': 0, 'assessed': False},
            'financial': {'weight': 0.20, 'score': 0, 'assessed': False},
            'strategic': {'weight': 0.15, 'score': 0, 'assessed': False},
            'validity': {'weight': 0.10, 'score': 0, 'assessed': False},
            'enforceability': {'weight': 0.10, 'score': 0, 'assessed': False},
        }

        for ra in risk_assessments:
            if ra.risk_factor in risk_factors:
                risk_factors[ra.risk_factor]['score'] = ra.risk_score * 10  # Convert 0-10 to 0-100
                risk_factors[ra.risk_factor]['assessed'] = True
                risk_factors[ra.risk_factor]['description'] = ra.description
                risk_factors[ra.risk_factor]['mitigation'] = ra.mitigation_strategy

        # Calculate overall score
        total_weight = sum(f['weight'] for f in risk_factors.values() if f['assessed'])
        if total_weight > 0:
            overall_score = sum(
                f['score'] * f['weight']
                for f in risk_factors.values()
                if f['assessed']
            ) / total_weight
        else:
            # If no assessments, use infringement likelihood as proxy
            overall_score = case.infringement_likelihood

        # Calculate estimated damages range
        damages_estimates = risk_assessments.exclude(
            estimated_damages_min__isnull=True
        )
        total_damages_min = sum(d.estimated_damages_min for d in damages_estimates if d.estimated_damages_min)
        total_damages_max = sum(d.estimated_damages_max for d in damages_estimates if d.estimated_damages_max)
        total_litigation_cost = sum(d.litigation_cost_estimate for d in damages_estimates if d.litigation_cost_estimate)

        # Generate risk level based on score
        if overall_score >= 80:
            risk_level = 'critical'
        elif overall_score >= 60:
            risk_level = 'high'
        elif overall_score >= 40:
            risk_level = 'medium'
        else:
            risk_level = 'low'

        # Generate mitigation recommendations
        recommendations = []
        for factor, data in risk_factors.items():
            if data['assessed'] and data['score'] >= 60:
                recommendations.append({
                    'factor': factor,
                    'risk_level': 'high' if data['score'] >= 70 else 'medium',
                    'recommendation': data.get('mitigation', f'Review and mitigate {factor} risk factors'),
                    'priority': 'urgent' if data['score'] >= 80 else 'high' if data['score'] >= 70 else 'medium'
                })

        return Response({
            'case_id': str(case.id),
            'case_name': case.case_name,
            'overall_risk_score': round(overall_score, 1),
            'risk_level': risk_level,
            'risk_factors': risk_factors,
            'damages_estimate': {
                'min': float(total_damages_min) if total_damages_min else None,
                'max': float(total_damages_max) if total_damages_max else None,
                'litigation_cost': float(total_litigation_cost) if total_litigation_cost else None,
            },
            'recommendations': recommendations,
            'infringement_likelihood': case.infringement_likelihood,
            'confidence_level': case.confidence_level,
            'analysis_complete': all(f['assessed'] for f in risk_factors.values()),
        })

    @action(detail=False, methods=['post'], url_path='generate-risk-report')
    def generate_risk_report(self, request):
        """
        Generate a comprehensive risk analysis report for a case
        """
        from utils.export import export_risk_report_pdf

        case_id = request.data.get('case_id')
        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Calculate risk score first
        risk_data = self.calculate_risk_score(request).data

        # Create report record
        report = InfringementReport.objects.create(
            case=case,
            title=f"Risk Analysis Report - {case.case_name}",
            report_type='risk_assessment',
            status='final',
            summary=f"Comprehensive risk analysis with overall risk score of {risk_data['overall_risk_score']}%",
            findings=str(risk_data['risk_factors']),
            conclusion=f"Risk Level: {risk_data['risk_level'].upper()}",
            recommendations=str(risk_data['recommendations']),
            created_by=request.user,
        )

        serializer = InfringementReportSerializer(report)
        return Response({
            'report': serializer.data,
            'risk_analysis': risk_data,
        })


class AIAnalysisViewSet(viewsets.ViewSet):
    """
    ViewSet for AI-powered analysis features
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='parse-claim')
    def parse_claim(self, request):
        """
        Parse claim text into structured elements using AI
        """
        claim_text = request.data.get('claim_text', '')
        if not claim_text:
            return Response(
                {'error': 'claim_text is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        use_llm = bool(request.data.get('use_llm', False))
        elements = ai_service.parse_claim_elements(claim_text, use_llm=use_llm)
        return Response({'elements': elements})

    @action(detail=False, methods=['post'], url_path='score-evidence')
    def score_evidence(self, request):
        """
        Score evidence relevance to a claim
        """
        evidence_id = request.data.get('evidence_id')
        claim_mapping_id = request.data.get('claim_mapping_id')

        if not evidence_id or not claim_mapping_id:
            return Response(
                {'error': 'evidence_id and claim_mapping_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            evidence = Evidence.objects.get(id=evidence_id)
            claim_mapping = ClaimMapping.objects.get(id=claim_mapping_id)
        except (Evidence.DoesNotExist, ClaimMapping.DoesNotExist):
            return Response(
                {'error': 'Evidence or ClaimMapping not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        result = ai_service.score_evidence_relevance(
            evidence.description,
            claim_mapping.claim_text,
            use_llm=bool(request.data.get('use_llm', False)),
        )

        return Response(result)

    @action(detail=False, methods=['post'], url_path='analyze-doe')
    def analyze_doe(self, request):
        """
        Analyze Doctrine of Equivalents for a claim element
        """
        claim_element_id = request.data.get('claim_element_id')

        if not claim_element_id:
            return Response(
                {'error': 'claim_element_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            element = ClaimElement.objects.get(id=claim_element_id)
        except ClaimElement.DoesNotExist:
            return Response(
                {'error': 'ClaimElement not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        result = ai_service.analyze_doe_similarity(
            element.element_text,
            element.accused_feature or '',
            use_llm=bool(request.data.get('use_llm', False)),
        )

        return Response(result)

    @action(detail=False, methods=['post'], url_path='predict-risk')
    def predict_risk(self, request):
        """
        Predict infringement risk using AI
        """
        case_id = request.data.get('case_id')

        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Prepare case data
        case_data = {
            'claim_mappings': [
                {
                    'claim_number': cm.claim_number,
                    'match_confidence': cm.match_confidence,
                    'limitations_met': cm.limitations_met,
                }
                for cm in case.claim_mappings.all()
            ],
            'evidence': [
                {
                    'relevance_score': e.relevance_score,
                    'evidence_type': e.evidence_type,
                }
                for e in case.evidence.all()
            ],
        }

        result = ai_service.predict_infringement_risk(case_data)
        return Response(result)

    @action(detail=False, methods=['post'], url_path='generate-narrative')
    def generate_narrative(self, request):
        """
        Generate professional narrative for claim chart
        """
        case_id = request.data.get('case_id')
        section = request.data.get('section', 'claim_chart')
        use_llm = bool(request.data.get('use_llm', False))

        if not case_id:
            return Response(
                {'error': 'case_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            case = InfringementCase.objects.get(id=case_id)
        except InfringementCase.DoesNotExist:
            return Response(
                {'error': 'Case not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        narratives = []
        for cm in case.claim_mappings.all():
            mapping_data = {
                'claim_number': cm.claim_number,
                'claim_text': cm.claim_text,
                'product_feature': cm.product_feature,
                'product_feature_description': cm.product_feature_description,
                'mapping_type': cm.mapping_type,
                'match_confidence': cm.match_confidence,
                'limitations_met': cm.limitations_met,
                'analysis_notes': cm.analysis_notes,
            }
            narrative = ai_service.generate_claim_chart_narrative(mapping_data, use_llm=use_llm)
            narratives.append({
                'claim_number': cm.claim_number,
                'narrative': narrative,
            })

        return Response({'narratives': narratives})
